import streamlit as st
import pandas as pd
import posixpath
import json
import re
import streamlit.components.v1 as components
from datetime import datetime, timezone
from io import BytesIO
import warnings
from openpyxl import load_workbook
from zoneinfo import ZoneInfo

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
)

from src.guard import (
    APP_TITLE,
    require_admin,
    hide_home_page_when_logged_in,
    hide_admin_page_for_non_admins,
    render_sidebar_header,
    render_logout_button,
)
from src.db import (
    list_users,
    get_user_by_username,
    count_admins,
    set_user_active,
    set_user_role,
    delete_user,
    # Scorecard DB helpers (Step 1)
    add_scorecard,
    list_scorecards,
    delete_scorecard_by_path,
    rebuild_blocks_from_fixtures_if_missing,
    list_blocks_with_fixtures,
    get_effective_block_state,
    get_block_open_at,
    get_current_block_number,
    set_block_override,
    clear_block_override,
    mark_block_scored,
    get_all_entries_for_block,
    upsert_block_player_points,
    upsert_block_user_points,
    list_block_user_points,
    get_block_prices,
    upsert_block_prices_from_dict,
    export_fantasy_backup_payload,
    restore_fantasy_from_backup_payload,
    wipe_all_fantasy_data,
    fantasy_has_state,
)
from src.auth import admin_reset_password, verify_password

from src.dropbox_api import (
    get_access_token,
    download_file,
    ensure_folder,
    upload_file,
    delete_path,
    list_folder,
)
from src.excel_io import load_league_workbook_from_bytes


st.set_page_config(page_title=f"{APP_TITLE} - Admin", layout="wide")

require_admin()
hide_home_page_when_logged_in()
hide_admin_page_for_non_admins()
render_sidebar_header()
render_logout_button()

def _format_last_login(series: pd.Series) -> pd.Series:
    """
    Format ISO timestamps to: DD Mon YYYY HH:MM
    Example: 16 Jan 2026 19:45
    """
    t = pd.to_datetime(series, errors="coerce", utc=True)
    return t.dt.strftime("%d %b %Y %H:%M")

def _get_secret(name: str) -> str:
    val = st.secrets.get(name, "")
    if not val:
        raise RuntimeError(f"Missing Streamlit secret: {name}")
    return str(val)


@st.cache_data(ttl=60, show_spinner=False)
def _load_workbook_fixture_results(app_key: str, app_secret: str, refresh_token: str, dropbox_path: str) -> pd.DataFrame:
    """
    Loads the league workbook from Dropbox and returns the fixture_results dataframe.
    Cached briefly to keep Admin UI responsive.
    """
    access_token = get_access_token(app_key, app_secret, refresh_token)
    xbytes = download_file(access_token, dropbox_path)
    data = load_league_workbook_from_bytes(xbytes)
    fixtures = data.fixture_results.copy()
    fixtures.columns = [str(c).strip() for c in fixtures.columns]
    return fixtures


@st.cache_data(ttl=60, show_spinner=False)
def _load_workbook_league_data(app_key: str, app_secret: str, refresh_token: str, dropbox_path: str) -> pd.DataFrame:
    """
    Loads the league workbook from Dropbox and returns the League_Data_Stats dataframe.
    Cached briefly to keep Admin UI responsive.
    """
    access_token = get_access_token(app_key, app_secret, refresh_token)
    xbytes = download_file(access_token, dropbox_path)
    data = load_league_workbook_from_bytes(xbytes)
    if data.league_data is None:
        return pd.DataFrame()
    league_data = data.league_data.copy()
    league_data.columns = [str(c).strip() for c in league_data.columns]
    return league_data


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

def _format_date_dd_mmm(series: pd.Series) -> pd.Series:
    dt = pd.to_datetime(series, errors="coerce", dayfirst=True)
    return dt.dt.strftime("%d-%b").fillna(series.astype(str))


def _format_time_ampm(series: pd.Series) -> pd.Series:
    formats = ("%H:%M", "%H:%M:%S", "%I %p", "%I:%M %p")

    def _format_one(val: object) -> str:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return str(val)
        raw = str(val).strip()
        if not raw or raw.lower() == "nan":
            return raw
        for fmt in formats:
            try:
                parsed = datetime.strptime(raw, fmt)
            except ValueError:
                continue
            hour = parsed.strftime("%I").lstrip("0") or "12"
            if parsed.minute == 0:
                return f"{hour} {parsed.strftime('%p')}"
            return f"{hour}:{parsed.strftime('%M')} {parsed.strftime('%p')}"
        return raw

    return series.apply(_format_one)

def _format_dt_dd_mmm_hhmm(dt_val: str | None) -> str | None:
    if dt_val is None:
        return None
    s = str(dt_val).strip()
    if not s:
        return s
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        return dt.strftime("%d-%b %H:%M")
    except Exception:
        return str(dt_val)


def _load_named_table_from_xlsm_bytes(xbytes: bytes, table_name: str) -> tuple[pd.DataFrame, bool]:
    wb = load_workbook(BytesIO(xbytes), data_only=True)
    for ws in wb.worksheets:
        if table_name in ws.tables:
            table = ws.tables[table_name]
            ref = table.ref
            cells = ws[ref]
            data = [[c.value for c in row] for row in cells]
            if not data or len(data) < 2:
                return pd.DataFrame(), True
            headers = [str(h).strip() if h is not None else "" for h in data[0]]
            rows = data[1:]
            df = pd.DataFrame(rows, columns=headers)
            blank_header_cols = [c for c in df.columns if str(c).strip() == ""]
            if blank_header_cols:
                df = df.drop(columns=blank_header_cols)
            return df, True
    return pd.DataFrame(), False


def _round_to_0_5(x: float) -> float:
    return round(x * 2) / 2


def _clamp(x: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, x))


def _safe_float(val: object) -> float | None:
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return None
        out = float(val)
    except Exception:
        return None
    if pd.isna(out):
        return None
    return out


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    return None


def _app_backup_folder(dropbox_file_path: str) -> str:
    app_folder = posixpath.dirname(dropbox_file_path.rstrip("/"))
    return posixpath.join(app_folder, "app_data")


def _fantasy_backup_path(dropbox_file_path: str) -> str:
    return posixpath.join(_app_backup_folder(dropbox_file_path), "fantasy_backup.json")


def _users_backup_path(dropbox_file_path: str) -> str:
    return posixpath.join(_app_backup_folder(dropbox_file_path), "users_backup.json")


def _fantasy_backup_to_dropbox(
    app_key: str, app_secret: str, refresh_token: str, backup_path: str
) -> None:
    access_token = get_access_token(app_key, app_secret, refresh_token)
    payload = export_fantasy_backup_payload()
    content = json.dumps(payload, indent=2).encode("utf-8")
    backup_folder = posixpath.dirname(backup_path)
    if backup_folder not in ("", "/"):
        ensure_folder(access_token, backup_folder)
    upload_file(access_token, backup_path, content, mode="overwrite", autorename=False)


def _fantasy_restore_from_dropbox_if_needed(
    app_key: str, app_secret: str, refresh_token: str, backup_path: str
) -> tuple[bool, str | None]:
    if fantasy_has_state():
        return False, None
    access_token = get_access_token(app_key, app_secret, refresh_token)
    try:
        raw = download_file(access_token, backup_path)
    except Exception:
        return False, None
    try:
        payload = json.loads(raw.decode("utf-8"))
        restore_fantasy_from_backup_payload(payload)
        return True, None
    except Exception as e:
        return False, str(e)

st.title("Admin")

tab_users, tab_scorecards, tab_fantasy_blocks = st.tabs(
    ["User Management", "Scorecard Management", "Fantasy"]
)

# =========================================================
# TAB 1: USER MANAGEMENT (existing functionality, unchanged)
# =========================================================
with tab_users:
    st.subheader("User Management")

    # Anchor placed above the user table (target for scroll)
    st.markdown('<div id="users_table_top"></div>', unsafe_allow_html=True)

    # One-time popup message after admin actions
    if st.session_state.get("admin_user_action_msg"):
        msg = st.session_state.pop("admin_user_action_msg")
        try:
            st.toast(msg, icon="✅")
        except Exception:
            st.success(msg)

    # One-time scroll back to the table/top
    if st.session_state.pop("admin_scroll_to_users", False):
        components.html(
            """
            <script>
            const el = window.parent.document.getElementById("users_table_top");
            if (el) { el.scrollIntoView({behavior: "smooth", block: "start"}); }
            else { window.parent.scrollTo({top: 0, behavior: "smooth"}); }
            </script>
            """,
            height=0,
        )
    
    # One-time popup message after admin actions
    if st.session_state.get("admin_user_action_msg"):
        msg = st.session_state.pop("admin_user_action_msg")

        # Prefer toast if available; fall back to success
        try:
            st.toast(msg, icon="✅")
        except Exception:
            st.success(msg)

    users = list_users()
    if not users:
        st.info("No users found.")
        st.stop()

    df = pd.DataFrame(users)
    df_display = df.copy()
    df_display["is_active"] = df_display["is_active"].map({1: "Active", 0: "Disabled"})

    # NEW: format last_login_at for display as "16 Jan 2026 19:45"
    if "last_login_at" in df_display.columns:
        t = pd.to_datetime(df_display["last_login_at"], errors="coerce", utc=True)
        df_display["last_login_at"] = t.dt.strftime("%d %b %Y %H:%M").fillna("Never")

    if "created_at" in df_display.columns:
        t = pd.to_datetime(df_display["created_at"], errors="coerce", utc=True)
        df_display["created_at"] = t.dt.strftime("%d %b %Y %H:%M").fillna("Never")

    # Keep the large users table collapsible to reduce scrolling for admins.
    with st.expander("All users", expanded=False):
        st.dataframe(
            df_display[
                ["username", "first_name", "last_name", "role", "is_active", "created_at", "last_login_at"]
            ],
            width="stretch",
            hide_index=True,
        )

    st.markdown("---")
    st.markdown("### Manage a user")

    # Build searchable labels starting with first name; add stable suffixes to keep labels unique.
    def _safe_str(value: object) -> str:
        if pd.isna(value):
            return ""
        return str(value).strip()

    def _user_sort_key(row: pd.Series) -> tuple:
        first = _safe_str(row.get("first_name"))
        last = _safe_str(row.get("last_name"))
        username = _safe_str(row.get("username"))
        return (
            1 if not first else 0,
            first.lower(),
            last.lower(),
            username.lower(),
        )

    def _base_user_label(row: pd.Series) -> str:
        first = _safe_str(row.get("first_name"))
        last = _safe_str(row.get("last_name"))
        username = _safe_str(row.get("username"))

        if first:
            display = first
        elif last:
            display = "(No first name)"
        else:
            display = username or "(No first name)"

        if last:
            display = f"{display} {last}"

        if username:
            return f"{display} ({username})"
        return display

    sorted_users = sorted(df.to_dict("records"), key=lambda row: _user_sort_key(pd.Series(row)))
    option_labels: list[str] = []
    label_to_username: dict[str, str] = {}
    label_counts: dict[str, int] = {}

    for row in sorted_users:
        series_row = pd.Series(row)
        base_label = _base_user_label(series_row)
        label_counts[base_label] = label_counts.get(base_label, 0) + 1

        suffix = ""
        if label_counts[base_label] > 1:
            suffix = _safe_str(row.get("id")) or _safe_str(row.get("user_id")) or _safe_str(row.get("created_at"))
            if suffix:
                suffix = f" [{suffix}]"
            else:
                suffix = f" [{label_counts[base_label]}]"

        label = f"{base_label}{suffix}"
        counter = 2
        while label in label_to_username:
            label = f"{base_label}{suffix} [{counter}]"
            counter += 1
        option_labels.append(label)
        label_to_username[label] = _safe_str(row.get("username"))

    # Avoid auto-selecting the first user to prevent accidental admin actions.
    # Enforce a "no selection" state until an admin explicitly chooses a user.
    try:
        selected_label = st.selectbox(
            "Select a user",
            option_labels,
            index=None,
            placeholder="Choose an option",
            key="admin_user_select",
        )
    except TypeError:
        labels_with_prompt = ["Choose an option"] + option_labels
        selected_label = st.selectbox(
            "Select a user",
            labels_with_prompt,
            index=0,
            key="admin_user_select",
        )

    # Don't call st.stop() here; it would prevent other admin tabs from rendering.
    if not selected_label or selected_label == "Choose an option":
        st.info("Type a first name to search, then select a user to manage.")
    else:
        # Keep all user-specific actions inside this block so nothing runs without a selection.
        selected_username = label_to_username[selected_label]

        selected = get_user_by_username(selected_username)
        if not selected:
            st.error("Selected user could not be found (it may have been deleted).")
            st.stop()

        selected_role = selected["role"]
        selected_active = bool(selected["is_active"])

        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**Name:** {selected['first_name']} {selected['last_name']}")
            st.write(f"**Username:** {selected['username']}")
        with col2:
            st.write(f"**Role:** {selected_role}")
            st.write(f"**Status:** {'Active' if selected_active else 'Disabled'}")

        admins_total = count_admins(active_only=False)

        def is_last_admin_target() -> bool:
            return selected_role == "admin" and admins_total == 1

        st.markdown("### Actions")

        with st.expander("Enable / Disable user", expanded=False):
            desired_active = st.radio(
                "Set account status",
                ["Active", "Disabled"],
                index=0 if selected_active else 1,
                horizontal=True,
                key="admin_user_status_radio",
            )
            make_active = desired_active == "Active"

            if st.button("Apply status change", width="stretch", key="admin_apply_status"):
                if not make_active and is_last_admin_target():
                    st.error("Blocked: You cannot disable the last remaining admin.")
                else:
                    set_user_active(selected_username, make_active)

                    status_txt = "Active" if make_active else "Disabled"
                    st.session_state["admin_user_action_msg"] = f"Updated '{selected_username}' status to {status_txt}."
                    st.session_state["admin_scroll_to_users"] = True

        with st.expander("Reset password", expanded=False):
            default_pw = str(st.secrets.get("DEFAULT_RESET_PASSWORD", "ResetMe123!"))
            st.write("Reset the selected user's password to the default reset password.")
            st.code(default_pw, language=None)
            if st.button(
                "Reset to default password", width="stretch", key="admin_reset_pw_btn"
            ):
                try:
                    admin_reset_password(selected_username, default_pw)

                    st.session_state["admin_user_action_msg"] = (
                        f"Password reset for '{selected_username}'. User will be prompted to change it on next login."
                    )
                    st.session_state["admin_scroll_to_users"] = True

                except Exception as e:
                    st.error(str(e))

        with st.expander("Change role", expanded=False):
            desired_role = st.selectbox(
                "Role",
                ["player", "admin"],
                index=0 if selected_role == "player" else 1,
                key="admin_role_select",
            )

            if st.button("Apply role change", width="stretch", key="admin_apply_role"):
                if desired_role == selected_role:
                    st.info("No change to apply.")
                else:
                    if selected_role == "admin" and desired_role == "player" and is_last_admin_target():
                        st.error("Blocked: You cannot demote the last remaining admin.")
                    else:
                        set_user_role(selected_username, desired_role)

                        st.session_state["admin_user_action_msg"] = (
                            f"Updated '{selected_username}' role to {desired_role}."
                        )
                        st.session_state["admin_scroll_to_users"] = True

        with st.expander("Delete user", expanded=False):
            st.warning("This permanently deletes the user account. This cannot be undone.")
            confirm = st.checkbox("I understand and want to delete this user", key="admin_delete_user_confirm")

            if st.button("Delete user", width="stretch", disabled=not confirm, key="admin_delete_user_btn"):
                if is_last_admin_target():
                    st.error("Blocked: You cannot delete the last remaining admin.")
                else:
                    delete_user(selected_username)

                    st.session_state["admin_user_action_msg"] = f"Deleted user '{selected_username}'."
                    st.session_state["admin_scroll_to_users"] = True


# =========================================================
# TAB 2: SCORECARD MANAGEMENT
# =========================================================
with tab_scorecards:
    st.subheader("Scorecard Management")

    # ---- Read Dropbox secrets (same as other pages) ----
    try:
        app_key = _get_secret("DROPBOX_APP_KEY")
        app_secret = _get_secret("DROPBOX_APP_SECRET")
        refresh_token = _get_secret("DROPBOX_REFRESH_TOKEN")
        dropbox_file_path = _get_secret("DROPBOX_FILE_PATH")  # Excel workbook path
    except Exception as e:
        st.error(str(e))
        st.stop()

    # Derive "app folder" from the workbook path, then add /scorecards
    app_folder = posixpath.dirname(dropbox_file_path.rstrip("/"))
    scorecards_root = posixpath.join(app_folder, "scorecards")

    # Load fixtures so admins can pick a MatchID confidently
    with st.spinner("Loading fixtures from Dropbox..."):
        try:
            fixtures_df = _load_workbook_fixture_results(app_key, app_secret, refresh_token, dropbox_file_path)
        except Exception as e:
            st.error(f"Failed to load fixtures from Dropbox: {e}")
            st.stop()

    if "MatchID" not in fixtures_df.columns:
        st.error("Cannot find required column 'MatchID' in fixtures. Please confirm it exists in the workbook.")
        st.stop()

    # Build a friendly label if columns exist
    cols = fixtures_df.columns.tolist()
    has_date = "Date" in cols
    has_time = "Time" in cols
    has_home = "Home Team" in cols
    has_away = "Away Team" in cols

    fixture_rows = fixtures_df.copy()

    # Format Admin fixture selector display to match main app
    if has_date:
        fixture_rows["Date"] = _format_date_dd_mmm(fixture_rows["Date"])
    if has_time:
        fixture_rows["Time"] = _format_time_ampm(fixture_rows["Time"])

    def _safe_str(v) -> str:
        if pd.isna(v):
            return ""
        return str(v).strip()
    
    def _clean_name_for_path(s: str) -> str:
        """
        Make a safe-ish filename component for Dropbox:
        - remove slashes
        - collapse whitespace
        - strip
        """
        s = (s or "").replace("/", "-").replace("\\", "-")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _next_named_filename(existing_names: set[str], base: str, ext: str) -> str:
        """
        For PDFs: base + ext (first), then base + ' 2' + ext, base + ' 3' + ext...
        For images (or any base that already includes a number placeholder in the base),
        you can pass base like '... Image' and it will find the next integer suffix.
        """
        base = base.strip()
        ext = ext if ext.startswith(".") else f".{ext}"

        # PDF-style: "Base.ext" then "Base 2.ext" ...
        # We will treat "Base.ext" as number 1.
        pattern = re.compile(rf"^{re.escape(base)}(?: (\d+))?{re.escape(ext)}$", re.IGNORECASE)

        max_n = 0
        for name in existing_names:
            m = pattern.match(name)
            if not m:
                continue
            n_txt = m.group(1)
            n = 1 if n_txt is None else int(n_txt)
            max_n = max(max_n, n)

        next_n = max_n + 1
        if next_n == 1:
            return f"{base}{ext}"
        return f"{base} {next_n}{ext}"

    def _next_image_filename(existing_names: set[str], base_prefix: str, ext: str) -> str:
        """
        Images: "BasePrefix 1.ext", "BasePrefix 2.ext"...
        """
        base_prefix = base_prefix.strip()
        ext = ext if ext.startswith(".") else f".{ext}"

        pattern = re.compile(rf"^{re.escape(base_prefix)} (\d+){re.escape(ext)}$", re.IGNORECASE)

        max_n = 0
        for name in existing_names:
            m = pattern.match(name)
            if not m:
                continue
            max_n = max(max_n, int(m.group(1)))

        return f"{base_prefix} {max_n + 1}{ext}"
    
    options = []
    option_to_match_id = {}

    for _, r in fixture_rows.iterrows():
        mid = _safe_str(r.get("MatchID"))
        if not mid:
            continue

        parts = [mid]
        if has_date:
            parts.append(_safe_str(r.get("Date")))
        if has_time:
            parts.append(_safe_str(r.get("Time")))
        if has_home and has_away:
            parts.append(f"{_safe_str(r.get('Home Team'))} vs {_safe_str(r.get('Away Team'))}")
        label = " — ".join([p for p in parts if p])

        options.append(label)
        option_to_match_id[label] = mid

    if not options:
        st.info("No fixtures with a valid MatchID were found.")
        st.stop()

    selected_label = st.selectbox("Select fixture", options, key="scorecard_match_select")
    match_id = option_to_match_id[selected_label]

    st.caption(f"Scorecard folder: {posixpath.join(scorecards_root, match_id)}")

    st.markdown("### Upload files")
        # Used to clear the file_uploader after successful upload
    if "scorecard_uploader_nonce" not in st.session_state:
        st.session_state["scorecard_uploader_nonce"] = 0

    uploader_key = f"scorecard_uploader_{match_id}_{st.session_state['scorecard_uploader_nonce']}"

    uploaded_files = st.file_uploader(
        "Upload scorecard PDFs or screenshots (you can select multiple files)",
        type=["pdf", "png", "jpg", "jpeg", "webp"],
        accept_multiple_files=True,
        key=uploader_key,
    )

    colu1, colu2 = st.columns([1, 2])
    with colu1:
        do_upload = st.button("Upload to Match", width="stretch", disabled=not uploaded_files, key="scorecard_upload_btn")
    with colu2:
        st.write("Uploads are appended (existing files are not removed). Filenames may be auto-renamed if duplicates exist.")

    if do_upload:
        try:
            access_token = get_access_token(app_key, app_secret, refresh_token)

            # Ensure /scorecards and /scorecards/<MatchID> exist
            ensure_folder(access_token, scorecards_root)
            match_folder = posixpath.join(scorecards_root, match_id)
            ensure_folder(access_token, match_folder)

            uploader_username = (st.session_state.get("user") or {}).get("username", "")

            # Pull Home/Away names for this MatchID (for renaming)
            fx = fixtures_df.copy()
            fx["MatchID"] = fx["MatchID"].astype(str).str.strip()
            fx_row = fx[fx["MatchID"] == str(match_id).strip()]

            home = ""
            away = ""
            if not fx_row.empty:
                r0 = fx_row.iloc[0]
                home = _safe_str(r0.get("Home Team"))
                away = _safe_str(r0.get("Away Team"))

            match_desc = _clean_name_for_path(f"{home} vs {away}".strip(" vs "))

            # Read existing Dropbox filenames in this match folder to pick next numbers consistently
            existing_entries = []
            try:
                existing_entries = list_folder(access_token, match_folder) or []
            except Exception:
                existing_entries = []

            existing_names = set()
            for e in existing_entries:
                nm = e.get("name")
                if nm:
                    existing_names.add(str(nm))

            for f in uploaded_files:
                original_name = f.name
                content = f.getvalue()

                # Extension handling
                ext = ""
                if "." in original_name:
                    ext = "." + original_name.split(".")[-1].lower().strip(".")
                else:
                    ext = ""

                is_pdf = ext == ".pdf"
                is_image = ext in [".png", ".jpg", ".jpeg", ".webp"]

                # Default fallback if we can't determine teams
                if not match_desc:
                    match_desc = _clean_name_for_path(f"Match {match_id}")

                # Build the new filename
                if is_pdf:
                    base = f"{match_desc} Scorecard"
                    new_name = _next_named_filename(existing_names, base=base, ext=ext)
                elif is_image:
                    # Normalise jpeg extension style to .jpeg if desired; keep original if you prefer.
                    if ext == ".jpg":
                        ext_use = ".jpeg"
                    else:
                        ext_use = ext

                    base_prefix = f"{match_desc} Scorecard Image"
                    new_name = _next_image_filename(existing_names, base_prefix=base_prefix, ext=ext_use)
                else:
                    # Unknown type: keep original name but still avoid collisions via Dropbox autorename
                    new_name = original_name

                # Ensure our set updates so multiple uploads in one click increment properly
                existing_names.add(new_name)

                dropbox_target_path = posixpath.join(match_folder, new_name)

                meta = upload_file(
                    access_token,
                    dropbox_target_path,
                    content,
                    mode="add",
                    autorename=True,  # still keep as a final backstop
                )

                dbx_path = meta.get("path_display") or meta.get("path_lower") or dropbox_target_path
                stored_name = meta.get("name") or new_name

                add_scorecard(
                    match_id=match_id,
                    file_name=stored_name,
                    dropbox_path=dbx_path,
                    uploaded_at=_utc_now_iso(),
                    uploaded_by=uploader_username,
                )

            st.success("Upload complete.")
            st.session_state["scorecard_uploader_nonce"] += 1

        except Exception as e:
            st.error(f"Upload failed: {e}")

        st.markdown("---")

    # =========================
    # Collapsible: Uploaded files
    # =========================
    with st.expander("Uploaded scorecards for this fixture", expanded=False):

        existing = list_scorecards(match_id)

        # Show in upload order (oldest first)
        existing = sorted(
            existing,
            key=lambda r: (str(r.get("uploaded_at") or ""), int(r.get("scorecard_id") or 0)),
        )
        # --- Reconcile SQLite records with what actually exists in Dropbox ---
        # If a file was deleted directly in Dropbox, remove the stale DB record
        # so the UI does not show phantom uploads.
        try:
            access_token = get_access_token(app_key, app_secret, refresh_token)
            match_folder = posixpath.join(scorecards_root, match_id)

            dbx_entries = list_folder(access_token, match_folder)

            # Normalise to a comparable set of paths.
            # Dropbox may return path_display and/or path_lower.
            dbx_paths = set()
            for e in dbx_entries:
                p_disp = e.get("path_display")
                p_low = e.get("path_lower")
                if p_disp:
                    dbx_paths.add(str(p_disp))
                    dbx_paths.add(str(p_disp).lower())
                if p_low:
                    dbx_paths.add(str(p_low))
                    dbx_paths.add(str(p_low).lower())

            stale = []
            for row in existing:
                p = str(row.get("dropbox_path", "") or "")
                if not p:
                    continue
                if p not in dbx_paths and p.lower() not in dbx_paths:
                    stale.append(p)

            # Auto-clean stale records (Dropbox file already gone)
            if stale:
                for p in stale:
                    delete_scorecard_by_path(p)

                # Re-load now-clean list for display
                existing = list_scorecards(match_id)

                # Show in upload order (oldest first)
                existing = sorted(
                    existing,
                    key=lambda r: (str(r.get("uploaded_at") or ""), int(r.get("scorecard_id") or 0)),
                )
                st.warning(
                    f"Cleaned up {len(stale)} stale scorecard record(s) (they were deleted directly in Dropbox)."
                )

        except Exception as e:
            # If Dropbox check fails, we still show DB list rather than breaking Admin.
            st.info(f"Dropbox cross-check unavailable (showing DB records only): {e}")

        if not existing:
            st.info("No scorecards uploaded yet for this Match.")
        else:
            for row in existing:
                fname = row.get("file_name", "")
                uploaded_at = row.get("uploaded_at", "")
                uploaded_by = row.get("uploaded_by", "")
                dbx_path = row.get("dropbox_path", "")
                scorecard_id = row.get("scorecard_id", "")

                with st.expander(f"{fname}", expanded=False):
                    st.write(f"**Uploaded at:** {uploaded_at}")
                    if uploaded_by:
                        st.write(f"**Uploaded by:** {uploaded_by}")
                    st.write(f"**Dropbox path:** `{dbx_path}`")

                    confirm_del = st.checkbox(
                        "I want to delete this file from Dropbox",
                        key=f"scorecard_del_confirm_{scorecard_id}",
                    )

                    if st.button(
                        "Delete file",
                        type="primary",
                        width="stretch",
                        disabled=not confirm_del,
                        key=f"scorecard_del_btn_{scorecard_id}",
                    ):
                        try:
                            access_token = get_access_token(app_key, app_secret, refresh_token)
                            delete_path(access_token, dbx_path)          # remove from Dropbox
                            delete_scorecard_by_path(dbx_path)           # remove from SQLite
                            st.success("Deleted.")
                        except Exception as e:
                            st.error(f"Delete failed: {e}")

    # =========================
    # Collapsible: Delete all
    # =========================
    with st.expander("Delete all files for this Match", expanded=False):
        st.warning(
            "This will permanently delete ALL uploaded scorecard files for this Match"
            "and remove their records from the database. This cannot be undone."
        )

        confirm_del_all = st.checkbox(
            "I understand and want to delete ALL scorecard files for this Match",
            key=f"scorecard_delete_all_confirm_{match_id}",
        )

        if st.button(
            "Delete ALL files for this Match",
            type="primary",
            width="stretch",
            disabled=not confirm_del_all,
            key=f"scorecard_delete_all_btn_{match_id}",
        ):
            try:
                access_token = get_access_token(app_key, app_secret, refresh_token)

                # 1) Delete all SQLite rows for this match
                existing_rows = list_scorecards(match_id)
                for row in existing_rows:
                    p = str(row.get("dropbox_path", "") or "")
                    if p:
                        delete_scorecard_by_path(p)

                # 2) Delete the entire Dropbox folder for this match (removes all files inside)
                match_folder = posixpath.join(scorecards_root, match_id)
                try:
                    delete_path(access_token, match_folder)
                except Exception:
                    pass  # Folder already gone is acceptable

                st.success("All scorecard files and database records for this Match have been deleted.")
            except Exception as e:
                st.error(f"Delete-all failed: {e}")


# =========================================================
# TAB 3: FANTASY BLOCKS
# =========================================================
with tab_fantasy_blocks:
    st.subheader("Fantasy")

    if st.session_state.get("admin_fantasy_msg"):
        msg = st.session_state.pop("admin_fantasy_msg")
        try:
            st.toast(msg, icon="✅")
        except Exception:
            st.success(msg)

    app_key = app_secret = refresh_token = dropbox_file_path = None
    backup_path = None
    try:
        app_key = _get_secret("DROPBOX_APP_KEY")
        app_secret = _get_secret("DROPBOX_APP_SECRET")
        refresh_token = _get_secret("DROPBOX_REFRESH_TOKEN")
        dropbox_file_path = _get_secret("DROPBOX_FILE_PATH")
        backup_path = _fantasy_backup_path(dropbox_file_path)
        users_backup_path = _users_backup_path(dropbox_file_path)
    except Exception:
        pass

    if backup_path and not st.session_state.get("fantasy_restore_attempted"):
        restored, err = _fantasy_restore_from_dropbox_if_needed(
            app_key, app_secret, refresh_token, backup_path
        )
        st.session_state["fantasy_restore_attempted"] = True
        if err:
            st.error(f"Fantasy restore failed: {err}")
        elif restored:
            st.info("Fantasy data restored from backup.")

    blocks = list_blocks_with_fixtures()
    if not blocks:
        # ---- Read Dropbox secrets (same as other pages) ----
        try:
            app_key = _get_secret("DROPBOX_APP_KEY")
            app_secret = _get_secret("DROPBOX_APP_SECRET")
            refresh_token = _get_secret("DROPBOX_REFRESH_TOKEN")
            dropbox_file_path = _get_secret("DROPBOX_FILE_PATH")  # Excel workbook path
        except Exception as e:
            st.error(str(e))
            st.stop()

        with st.spinner("Loading fixtures from Dropbox..."):
            try:
                fixtures_df = _load_workbook_fixture_results(app_key, app_secret, refresh_token, dropbox_file_path)
            except Exception as e:
                st.error(f"Failed to load fixtures from Dropbox: {e}")
                st.stop()

        created_blocks = rebuild_blocks_from_fixtures_if_missing(fixtures_df)
        if created_blocks:
            st.success(f"Created {created_blocks} fantasy block(s) from fixtures.")

        blocks = list_blocks_with_fixtures()

    london_now = datetime.now(ZoneInfo("Europe/London"))
    london_now_iso = london_now.isoformat()
    if not blocks:
        st.info("No fantasy blocks found.")
    else:
        rows = []
        for b in blocks:
            open_at = get_block_open_at(b.get("block_number"), london_now)
            fixtures_list = []
            for fx in b.get("fixtures", []):
                match_id = fx.get("match_id", "")
                start_at = fx.get("start_at", "")
                fixtures_list.append(f"{match_id} ({_format_dt_dd_mmm_hhmm(start_at)})")

            rows.append(
                {
                    "block_number": b.get("block_number"),
                    "first_start_at": _format_dt_dd_mmm_hhmm(b.get("first_start_at")),
                    "lock_at": _format_dt_dd_mmm_hhmm(b.get("lock_at")),
                    "open_at": _format_dt_dd_mmm_hhmm(open_at),
                    "scored_at": _format_dt_dd_mmm_hhmm(b.get("scored_at")),
                    "override_state": b.get("override_state"),
                    "override_until": _format_dt_dd_mmm_hhmm(b.get("override_until")),
                    "effective_state": get_effective_block_state(b.get("block_number"), london_now),
                    "fixtures": ", ".join(fixtures_list),
                }
            )

        st.dataframe(
            pd.DataFrame(rows),
            width="stretch",
            hide_index=True,
        )

    st.markdown("---")
    st.markdown("### Current Block Controls")

    current_block = get_current_block_number()
    if current_block is None:
        st.info("All blocks are scored.")
    else:
        current_state = get_effective_block_state(current_block, london_now)
        current_scored_at = None
        for b in blocks:
            if int(b.get("block_number") or 0) == int(current_block):
                current_scored_at = b.get("scored_at")
                break
        st.write(f"**Current block:** {current_block} (state: {current_state})")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            if st.button(
                "Lock now",
                width="stretch",
                key="fantasy_block_lock_now",
                disabled=current_state == "SCORED",
            ):
                set_block_override(current_block, "LOCKED", override_until=None)
                if backup_path and app_key and app_secret and refresh_token:
                    try:
                        _fantasy_backup_to_dropbox(
                            app_key, app_secret, refresh_token, backup_path
                        )
                    except Exception as e:
                        st.error(f"Fantasy backup failed: {e}")
                st.session_state["admin_fantasy_msg"] = f"Block {current_block} locked."
            if st.button(
                "Backup fantasy now",
                width="stretch",
                key="fantasy_backup_now",
            ):
                if backup_path and app_key and app_secret and refresh_token:
                    try:
                        _fantasy_backup_to_dropbox(
                            app_key, app_secret, refresh_token, backup_path
                        )
                        st.success("Fantasy backup uploaded.")
                        try:
                            access_token = get_access_token(app_key, app_secret, refresh_token)
                            raw = download_file(access_token, backup_path)
                            json.loads(raw.decode("utf-8"))
                            st.success("Backup verified (download + parse OK).")
                        except Exception as e:
                            st.warning(f"Backup verification failed: {e}")
                    except Exception as e:
                        st.error(f"Fantasy backup failed: {e}")
                else:
                    st.error("Dropbox configuration is missing.")
        with col2:
            if st.button(
                "Unlock now",
                width="stretch",
                key="fantasy_block_unlock_now",
                disabled=current_state == "SCORED",
            ):
                set_block_override(current_block, "OPEN", override_until=None)
                if backup_path and app_key and app_secret and refresh_token:
                    try:
                        _fantasy_backup_to_dropbox(
                            app_key, app_secret, refresh_token, backup_path
                        )
                    except Exception as e:
                        st.error(f"Fantasy backup failed: {e}")
                st.session_state["admin_fantasy_msg"] = f"Block {current_block} unlocked."
        with col3:
            if st.button(
                "Clear override",
                width="stretch",
                key="fantasy_block_clear_override",
                disabled=current_state == "SCORED",
            ):
                clear_block_override(current_block)
                if backup_path and app_key and app_secret and refresh_token:
                    try:
                        _fantasy_backup_to_dropbox(
                            app_key, app_secret, refresh_token, backup_path
                        )
                    except Exception as e:
                        st.error(f"Fantasy backup failed: {e}")
                st.session_state["admin_fantasy_msg"] = f"Block {current_block} override cleared."
        with col4:
            confirm_key = f"fantasy_block_confirm_week_{current_block}"
            confirm_text = f"I confirm Week {current_block} Stats is complete"
            confirm_ok = st.checkbox(
                confirm_text,
                value=False,
                key=confirm_key,
                disabled=current_state == "SCORED",
            )
            if st.button(
                "Score block (stats entered)",
                width="stretch",
                key="fantasy_block_score_now",
                disabled=(current_state == "SCORED" or not confirm_ok),
            ):
                try:
                    if current_scored_at:
                        st.warning(f"Block {current_block} is already scored.")
                        st.stop()
                    app_key = _get_secret("DROPBOX_APP_KEY")
                    app_secret = _get_secret("DROPBOX_APP_SECRET")
                    refresh_token = _get_secret("DROPBOX_REFRESH_TOKEN")
                    dropbox_file_path = _get_secret("DROPBOX_FILE_PATH")
                except Exception as e:
                    st.error(str(e))
                    st.stop()

                with st.spinner("Scoring block from Dropbox..."):
                    access_token = get_access_token(app_key, app_secret, refresh_token)
                    xbytes = download_file(access_token, dropbox_file_path)
                    table_name = f"Week{current_block}Stats"
                    week_df, table_found = _load_named_table_from_xlsm_bytes(xbytes, table_name)

                    if not table_found:
                        st.error(f"Missing table '{table_name}' in the workbook.")
                        st.stop()
                    if week_df is None or week_df.empty:
                        st.error(f"Table '{table_name}' is empty.")
                        st.stop()

                    week_df.columns = [str(c).strip() for c in week_df.columns]
                    if "PlayerID" not in week_df.columns or "Fantasy Points" not in week_df.columns:
                        st.error("Week table must include 'PlayerID' and 'Fantasy Points' columns.")
                        st.stop()

                    pts = week_df[["PlayerID", "Fantasy Points"]].copy()
                    pts["PlayerID"] = pts["PlayerID"].astype(str).str.strip()
                    pts["Fantasy Points"] = pd.to_numeric(pts["Fantasy Points"], errors="coerce")
                    pts = pts[(pts["PlayerID"] != "") & (pts["Fantasy Points"].notna())]
                    points_by_player = (
                        pts.groupby("PlayerID")["Fantasy Points"].sum().to_dict()
                        if not pts.empty
                        else {}
                    )

                    if not points_by_player:
                        st.error(
                            f"Week{current_block}Stats contains no usable stats (PlayerID/Fantasy Points). Block was NOT scored."
                        )
                        st.stop()

                    entries = get_all_entries_for_block(current_block)
                    user_points: dict[int, float] = {}
                    if not entries:
                        st.info("No teams submitted for this block.")

                    for entry in entries:
                        entry_players = entry.get("entry_players", [])
                        starting_ids = [
                            str(r["player_id"])
                            for r in entry_players
                            if int(r.get("is_starting") or 0) == 1
                        ]
                        bench1_id = ""
                        bench2_id = ""
                        captain_id = ""
                        vice_id = ""
                        for r in entry_players:
                            if r.get("bench_order") == 1:
                                bench1_id = str(r["player_id"])
                            elif r.get("bench_order") == 2:
                                bench2_id = str(r["player_id"])
                            if int(r.get("is_captain") or 0) == 1:
                                captain_id = str(r["player_id"])
                            if int(r.get("is_vice_captain") or 0) == 1:
                                vice_id = str(r["player_id"])

                        bench_queue = []
                        for bid in [bench1_id, bench2_id]:
                            if bid and bid in points_by_player:
                                bench_queue.append(bid)

                        final_on_field = []
                        for sid in starting_ids:
                            if sid in points_by_player:
                                final_on_field.append(sid)
                            else:
                                if bench_queue:
                                    final_on_field.append(bench_queue.pop(0))
                                else:
                                    final_on_field.append(None)

                        total = 0.0
                        for pid in final_on_field:
                            if pid and pid in points_by_player:
                                total += float(points_by_player.get(pid, 0.0))

                        if captain_id in final_on_field and captain_id in points_by_player:
                            total += float(points_by_player.get(captain_id, 0.0))
                        if vice_id in final_on_field and vice_id in points_by_player and vice_id != captain_id:
                            total += float(points_by_player.get(vice_id, 0.0)) * 0.5

                        user_points[int(entry["user_id"])] = float(total)

                    upsert_block_player_points(current_block, points_by_player)
                    upsert_block_user_points(current_block, user_points, london_now_iso)
                    mark_block_scored(current_block, london_now_iso)

                    league_data_df = pd.DataFrame()
                    if app_key and app_secret and refresh_token and dropbox_file_path:
                        try:
                            league_data_df = _load_workbook_league_data(
                                app_key, app_secret, refresh_token, dropbox_file_path
                            )
                        except Exception:
                            league_data_df = pd.DataFrame()

                    played_players = list(points_by_player.keys())
                    current_prices = get_block_prices(current_block)
                    if not current_prices:
                        current_prices = {pid: 7.5 for pid in played_players}

                    if played_players:
                        pts_series = pd.Series([float(points_by_player[p]) for p in played_players])
                        median = float(pts_series.median())
                        q25 = float(pts_series.quantile(0.25))
                        q75 = float(pts_series.quantile(0.75))
                        iqr = q75 - q25
                    else:
                        median = 0.0
                        iqr = 0.0

                    appm_by_pid: dict[str, float] = {}
                    matches_by_pid: dict[str, float] = {}
                    use_appm = False
                    if not league_data_df.empty:
                        tmp = league_data_df.copy()
                        tmp.columns = [str(c).strip() for c in tmp.columns]
                        pid_col = _find_col(tmp, ["PlayerID", "Player Id", "Player ID"])
                        appm_col = _find_col(
                            tmp,
                            [
                                "Ave Points Per Match",
                                "Avg Points Per Match",
                                "Average Points Per Match",
                                "Ave Pts Per Match",
                                "Avg Pts Per Match",
                            ],
                        )
                        matches_col = _find_col(
                            tmp,
                            ["Matches Played", "Match Played", "Games Played", "Played"],
                        )
                        points_col = _find_col(tmp, ["Fantasy Points", "Total Points", "Points", "Pts"])

                        if pid_col and matches_col:
                            if appm_col:
                                tmp[appm_col] = pd.to_numeric(tmp[appm_col], errors="coerce")
                            if matches_col:
                                tmp[matches_col] = pd.to_numeric(tmp[matches_col], errors="coerce")
                            if points_col:
                                tmp[points_col] = pd.to_numeric(tmp[points_col], errors="coerce")

                            for _, row in tmp.iterrows():
                                pid_val = str(row.get(pid_col) or "").strip()
                                if not pid_val:
                                    continue
                                matches = _safe_float(row.get(matches_col))
                                if matches is None or matches <= 0:
                                    continue

                                appm = _safe_float(row.get(appm_col)) if appm_col else None
                                if appm is None and points_col:
                                    pts = _safe_float(row.get(points_col))
                                    if pts is not None:
                                        appm = pts / matches
                                if appm is None:
                                    continue

                                appm_by_pid[pid_val] = float(appm)
                                matches_by_pid[pid_val] = float(matches)

                            if appm_by_pid:
                                use_appm = True

                    if use_appm and played_players:
                        appm_vals = [
                            float(appm_by_pid[p])
                            for p in played_players
                            if p in appm_by_pid
                        ]
                        if appm_vals:
                            appm_series = pd.Series(appm_vals)
                            median_appm = float(appm_series.median())
                            q25_appm = float(appm_series.quantile(0.25))
                            q75_appm = float(appm_series.quantile(0.75))
                            iqr_appm = q75_appm - q25_appm
                        else:
                            median_appm = 0.0
                            iqr_appm = 0.0
                    else:
                        median_appm = 0.0
                        iqr_appm = 0.0

                    k = 0.5
                    next_prices: dict[str, float] = {}
                    for pid in played_players:
                        current_price = float(current_prices.get(pid, 7.5))
                        delta_raw = 0.0
                        if use_appm and pid in appm_by_pid:
                            matches = matches_by_pid.get(pid, 0.0)
                            new_appm = float(appm_by_pid[pid])
                            match_points = float(points_by_player.get(pid, 0.0))
                            if matches and matches > 1:
                                old_total = (new_appm * matches) - match_points
                                old_appm = old_total / (matches - 1)
                            else:
                                old_appm = 0.0
                            delta_appm = new_appm - old_appm
                            denom = max(iqr_appm, 1.0)
                            delta_raw = k * delta_appm / denom
                        elif pid in points_by_player:
                            denom = max(iqr, 1.0)
                            delta_raw = k * (float(points_by_player[pid]) - median) / denom
                        delta_capped = _clamp(delta_raw, -1.0, 1.0)
                        delta = _round_to_0_5(delta_capped)
                        price_next = _clamp(_round_to_0_5(current_price + delta), 5.0, 10.0)
                        next_prices[pid] = float(price_next)

                    if next_prices:
                        upsert_block_prices_from_dict(current_block + 1, next_prices)

                    leaderboard = list_block_user_points(current_block)
                    if leaderboard:
                        st.session_state["admin_fantasy_leaderboard"] = leaderboard

                    if backup_path and app_key and app_secret and refresh_token:
                        try:
                            _fantasy_backup_to_dropbox(
                                app_key, app_secret, refresh_token, backup_path
                            )
                        except Exception as e:
                            st.error(f"Fantasy backup failed: {e}")

                    st.session_state["admin_fantasy_msg"] = f"Block {current_block} scored."

    if st.session_state.get("admin_fantasy_leaderboard"):
        st.markdown("### Block Leaderboard")
        lb_rows = st.session_state["admin_fantasy_leaderboard"]
        display_rows = []
        for r in lb_rows:
            row = dict(r)
            if "calculated_at" in row:
                row["calculated_at"] = _format_dt_dd_mmm_hhmm(row.get("calculated_at"))
            display_rows.append(row)
        st.dataframe(
            pd.DataFrame(display_rows),
            width="stretch",
            hide_index=True,
        )

    st.markdown("---")
    st.subheader("Reset Fantasy League")
    st.warning(
        "This will permanently delete ALL fantasy data (blocks, entries, prices, and results). "
        "This cannot be undone."
    )
    confirm_reset = st.checkbox(
        "I understand this will delete all fantasy data",
        key="fantasy_reset_confirm",
    )
    reset_pw = st.text_input(
        "Re-enter your password to confirm",
        type="password",
        key="fantasy_reset_password",
    )
    if st.button("Reset Fantasy League", type="primary", width="stretch"):
        if not confirm_reset:
            st.error("Please confirm you understand the reset.")
        else:
            current_user = st.session_state.get("user") or {}
            username = str(current_user.get("username") or "").strip()
            if not username:
                st.error("Unable to identify current user.")
            else:
                user_row = get_user_by_username(username)
                if not user_row or not verify_password(reset_pw, user_row.get("password_hash", "")):
                    st.error("Password confirmation failed.")
                else:
                    wipe_all_fantasy_data()
                    if backup_path and app_key and app_secret and refresh_token:
                        try:
                            _fantasy_backup_to_dropbox(
                                app_key, app_secret, refresh_token, backup_path
                            )
                        except Exception:
                            pass
                    for k in list(st.session_state.keys()):
                        if k.startswith("fantasy_editing_block_"):
                            st.session_state.pop(k, None)
                    st.session_state.pop("fantasy_last_block_number", None)
                    st.session_state.pop("fantasy_restore_attempted", None)
                    st.session_state.pop("admin_fantasy_leaderboard", None)
                    st.success("Fantasy league reset.")
