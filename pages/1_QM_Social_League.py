"""QM Social League page: fixtures, stats, scorecards, and top performers views."""

from datetime import datetime
import logging
from io import BytesIO

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

from src.guard import (
    APP_TITLE,
    require_login,
    hide_home_page_when_logged_in,
    hide_admin_page_for_non_admins,
    render_sidebar_header,
    render_logout_button,
)
from src.dropbox_api import get_access_token, download_file, get_temporary_link
from src.excel_io import load_league_workbook_from_bytes, load_named_table_from_bytes
from src.db import list_scorecards, list_scorecard_match_ids

st.set_page_config(page_title=f"{APP_TITLE} - QM Social League", layout="wide")

require_login()
hide_home_page_when_logged_in()
hide_admin_page_for_non_admins()
render_sidebar_header()
render_logout_button()

st.title("QM Social League")
logger = logging.getLogger(__name__)
SEASON_SHEETS = {
    "Current Season": "League_Data",
    "Sem A 25/26": "Sem_A_25-26_Stats",
    "Sem B 24/25": "Sem_B_24-25_Stats",
}


def _get_secret(name: str) -> str:
    val = st.secrets.get(name, "")
    if not val:
        raise RuntimeError(f"Missing Streamlit secret: {name}")
    return str(val)


@st.cache_data(ttl=60, show_spinner=False)
def _load_from_dropbox(app_key: str, app_secret: str, refresh_token: str, dropbox_path: str):
    access_token = get_access_token(app_key, app_secret, refresh_token)
    xbytes = download_file(access_token, dropbox_path)
    return load_league_workbook_from_bytes(xbytes)


@st.cache_data(ttl=60, show_spinner=False)
def _download_workbook_bytes(app_key: str, app_secret: str, refresh_token: str, dropbox_path: str) -> bytes:
    access_token = get_access_token(app_key, app_secret, refresh_token)
    return download_file(access_token, dropbox_path)


@st.cache_data(ttl=300, show_spinner=False)
def _download_scorecard_bytes(app_key: str, app_secret: str, refresh_token: str, dropbox_path: str) -> bytes:
    """Download a scorecard file from Dropbox (cached briefly for UX)."""
    access_token = get_access_token(app_key, app_secret, refresh_token)
    return download_file(access_token, dropbox_path)


@st.cache_data(ttl=60, show_spinner=False)
def _get_temp_link(app_key: str, app_secret: str, refresh_token: str, dropbox_path: str) -> str:
    access_token = get_access_token(app_key, app_secret, refresh_token)
    return get_temporary_link(access_token, dropbox_path)


@st.cache_data(ttl=60, show_spinner=False)
def _match_has_scorecards(match_id: str) -> bool:
    """Fast check to filter the fixture selector to only fixtures with uploads."""
    return len(list_scorecards(match_id)) > 0


def _format_date_dd_mmm(series: pd.Series) -> pd.Series:
    dt = pd.to_datetime(series, errors="coerce", dayfirst=True)
    return dt.dt.strftime("%d-%b").fillna(series.astype(str))


def _format_time_ampm(series: pd.Series) -> pd.Series:
    formats = ("%H:%M", "%H:%M:%S", "%I %p", "%I:%M %p")

    def _format_one(val: object) -> str:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return str(val)
        raw = str(val).strip()
        if not raw or raw.lower() == "nan":
            return raw
        for fmt in formats:
            try:
                parsed = datetime.strptime(raw, fmt)
            except ValueError:
                continue
            hour = parsed.strftime("%I").lstrip("0") or "12"
            if parsed.minute == 0:
                return f"{hour} {parsed.strftime('%p')}"
            return f"{hour}:{parsed.strftime('%M')} {parsed.strftime('%p')}"
        return raw

    return series.apply(_format_one)


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    return None


def _find_col_case_insensitive(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lookup = {str(c).strip().casefold(): c for c in df.columns}
    for c in candidates:
        found = lookup.get(str(c).strip().casefold())
        if found is not None:
            return str(found)
    return None


def _filter_valid_players(df: pd.DataFrame) -> pd.DataFrame:
    """
    Keep only real player rows.
    Drops rows where:
    - PlayerID is blank/null or equals "missing" (case-insensitive), OR
    - Player name is blank/null.
    If one column is absent, filter using whichever identity column is available.
    """
    if df is None or df.empty:
        return df

    out = df.copy()
    id_col = _find_col_case_insensitive(out, ["PlayerID", "player_id", "Player Id", "Player ID"])
    name_col = _find_col_case_insensitive(out, ["Player", "Player Name", "Name"])

    invalid_mask = pd.Series(False, index=out.index)

    if id_col and id_col in out.columns:
        id_raw = out[id_col]
        id_str = id_raw.astype(str).str.strip()
        id_invalid = id_raw.isna() | (id_str == "") | (id_str.str.casefold() == "missing")
        invalid_mask = invalid_mask | id_invalid

    if name_col and name_col in out.columns:
        name_raw = out[name_col]
        name_str = name_raw.astype(str).str.strip()
        name_invalid = name_raw.isna() | (name_str == "")
        invalid_mask = invalid_mask | name_invalid

    if not (id_col or name_col):
        return out

    return out[~invalid_mask].copy()


def _normalize_playerid_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    Streamlit uses Arrow; mixed types in object columns (int + str) can cause ArrowInvalid.
    Force PlayerID-like columns to clean strings.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    for c in ["PlayerID", "Player Id", "Player ID"]:
        if c in out.columns:
            out[c] = out[c].astype(str).str.strip()
            out[c] = out[c].replace({"nan": "", "None": "", "NaT": ""})
    return out


ALL = "All"


def _as_list(v):
    if v is None:
        return []
    if isinstance(v, str):
        return [v]
    try:
        return list(v)
    except Exception:
        return []


def enforce_all_exclusive(key: str, all_token: str = ALL) -> None:
    sel = _as_list(st.session_state.get(key))
    prev_key = f"__prev_{key}"
    prev = _as_list(st.session_state.get(prev_key))

    if all_token in sel:
        if all_token not in prev:
            st.session_state[key] = [all_token]
            return
        if len(sel) > 1:
            st.session_state[key] = [x for x in sel if x != all_token]
            return


def _resolved_selection(
    widget_key: str,
    real_options: list[str],
    default_list: list[str],
    all_label: str = ALL,
) -> list[str]:
    sel = _as_list(st.session_state.get(widget_key, default_list))
    if not sel:
        return []
    if all_label in sel:
        return real_options[:]
    return [x for x in sel if x in real_options]


def _init_or_sanitize_multiselect_state_allow_empty(key: str, options: list[str], defaults: list[str]) -> None:
    """
    - First load: set defaults (only those in options).
    - Later loads: remove invalid selections only.
    - Allow user to clear all selections (empty remains empty).
    """
    if key not in st.session_state:
        st.session_state[key] = [c for c in defaults if c in options]
        return

    current = st.session_state.get(key, [])
    if current is None:
        current = []
    st.session_state[key] = [c for c in current if c in options]


def _extract_teams_df(excel_result) -> pd.DataFrame | None:
    teams_df = getattr(excel_result, "teams_table", None)
    if teams_df is None:
        teams_df = getattr(excel_result, "teams", None)
    if teams_df is None:
        teams_df = getattr(excel_result, "teams_data", None)
    return teams_df


def parse_top_performers_grid(df_grid: pd.DataFrame) -> dict:
    grid = df_grid.fillna("").astype(str).values.tolist()
    section_titles = ("Top Performers", "Top Performers - conditional")
    categories = ("Batting", "Bowling", "Fielding")

    parsed: dict[str, dict[str, list[dict[str, str]]]] = {}
    current_section: str | None = None
    current_category: str | None = None

    def _get_cell(row: list[str], idx: int) -> str:
        if idx < 0 or idx >= len(row):
            return ""
        return str(row[idx]).strip()

    def _ensure_section(name: str) -> None:
        if name not in parsed:
            parsed[name] = {
                "Batting": [],
                "Bowling": [],
                "Fielding": [],
                "Meta": [],
            }

    for row in grid:
        cells = [str(c).strip() for c in row]
        if not any(cells):
            continue

        section_hit = next((c for c in cells if c in section_titles), None)
        if section_hit:
            current_section = section_hit
            _ensure_section(current_section)
            current_category = None
            continue

        if not current_section:
            continue

        c0 = _get_cell(row, 0)
        if c0 in categories:
            current_category = c0

        metric = _get_cell(row, 1)
        value = _get_cell(row, 2)
        player = _get_cell(row, 3)

        if not metric:
            continue

        if metric.lower().startswith("minimum "):
            meta_val = player or value
            parsed[current_section]["Meta"].append(
                {"label": metric.rstrip(":"), "value": meta_val}
            )
            continue

        if current_category in categories:
            parsed[current_section][current_category].append(
                {"metric": metric, "value": value, "player": player}
            )

    return parsed


def _format_top_performer_value(metric: str, value: str) -> str:
    metric_l = str(metric or "").strip().lower()
    raw = str(value or "").strip()
    if not raw or "-" in raw:
        return raw
    try:
        num = float(raw)
    except Exception:
        return raw
    if any(k in metric_l for k in ("average", "economy", "strike rate", "avg")):
        return f"{num:.2f}".rstrip("0").rstrip(".")
    return raw


def render_top_performers_streamlit(parsed: dict) -> None:
    section_names = ["Top Performers", "Top Performers - conditional"]
    available = [s for s in section_names if s in parsed]
    if not available:
        st.info("Top performers are not available yet.")
        return

    for section_idx, section_name in enumerate(available):
        section = parsed.get(section_name) or {}
        st.subheader(section_name)

        meta_items = section.get("Meta") or []
        if meta_items:
            meta_line = " • ".join(
                f"{str(m.get('label') or '').strip()}: {str(m.get('value') or '').strip()}"
                for m in meta_items
                if str(m.get("label") or "").strip()
            )
            if meta_line:
                st.caption(meta_line)

        categories = ["Batting", "Bowling", "Fielding"]
        if section_name == "Top Performers - conditional":
            categories = ["Batting", "Bowling"]
        tabs = st.tabs(categories)
        for tab, category in zip(tabs, categories):
            with tab:
                rows = section.get(category) or []
                if not rows:
                    st.info(f"No {category.lower()} top performers in this section.")
                    continue
                display_rows = []
                for r in rows:
                    metric = str(r.get("metric") or "").strip()
                    val = _format_top_performer_value(metric, str(r.get("value") or "").strip())
                    player = str(r.get("player") or "").strip()
                    display_rows.append({"Metric": metric, "Value": val, "Player": player})
                st.dataframe(
                    pd.DataFrame(display_rows, columns=["Metric", "Value", "Player"]),
                    use_container_width=True,
                    hide_index=True,
                )

        if section_idx < len(available) - 1:
            st.divider()


@st.cache_data(ttl=300, show_spinner=False)
def discover_seasons() -> dict[str, str]:
    return dict(SEASON_SHEETS)


@st.cache_data(ttl=300, show_spinner=False)
def _worksheet_to_df(workbook_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    try:
        wb = load_workbook(BytesIO(workbook_bytes), data_only=True)
    except Exception as exc:
        logger.warning("Failed to open workbook bytes for sheet '%s': %s", sheet_name, exc)
        return pd.DataFrame()

    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found in workbook", sheet_name)
        return pd.DataFrame()

    ws = wb[sheet_name]
    rows: list[list[object]] = []
    if ws.tables:
        table_name = "League_Data_Stats" if "League_Data_Stats" in ws.tables else next(iter(ws.tables.keys()))
        ref = ws.tables[table_name].ref
        for row in ws[ref]:
            rows.append([cell.value for cell in row])
    else:
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))

    if len(rows) < 2:
        return pd.DataFrame()

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = rows[1:]
    df = pd.DataFrame(data_rows, columns=headers)
    blank_cols = [c for c in df.columns if str(c).strip() == ""]
    if blank_cols:
        df = df.drop(columns=blank_cols)
    if not df.empty:
        df = df.dropna(axis=1, how="all")
        df = df.dropna(axis=0, how="all")
    return df


@st.cache_data(ttl=300, show_spinner=False)
def load_stats_for_sheet(workbook_bytes: bytes, sheet_name: str) -> pd.DataFrame:
    league = _worksheet_to_df(workbook_bytes, sheet_name)
    if league is None or league.empty:
        return pd.DataFrame()
    league.columns = [str(c).strip() for c in league.columns]
    if _find_col(league, ["Name", "Player", "Player Name"]) is None:
        logger.warning("Skipping sheet '%s': missing player name column", sheet_name)
        return pd.DataFrame()
    league = _filter_valid_players(league)
    return league if league is not None else pd.DataFrame()


@st.cache_data(ttl=300, show_spinner=False)
def load_combined_stats_table(workbook_bytes: bytes, table_name: str = "Combined_Stats") -> pd.DataFrame:
    df = load_named_table_from_bytes(workbook_bytes, table_name, drop_empty_columns=True)
    if df is None or df.empty:
        return pd.DataFrame()
    df.columns = [str(c).strip() for c in df.columns]
    df = _filter_valid_players(df)
    return df if df is not None else pd.DataFrame()


def render_player_stats_ui(
    df: pd.DataFrame,
    enable_team_filter: bool,
    current_season: bool,
    teams_df: pd.DataFrame | None = None,
    season_label: str | None = None,
) -> None:
    league = df.copy()
    league.columns = [str(c).strip() for c in league.columns]

    team_id_col_league = _find_col(league, ["TeamID", "Team Id", "Team ID"])
    name_col = _find_col(league, ["Name", "Player", "Player Name"])
    if not name_col:
        st.info("No player stats found yet (player name column is missing).")
        return

    team_id_to_name: dict[str, str] = {}
    team_name_to_id: dict[str, str] = {}
    if teams_df is not None and not teams_df.empty:
        teams = teams_df.copy()
        teams.columns = [str(c).strip() for c in teams.columns]
        team_id_col_teams = _find_col(teams, ["TeamID", "Team Id", "Team ID"])
        team_name_col_teams = _find_col(teams, ["Team Names", "Team Name", "Team"])
        if team_id_col_teams and team_name_col_teams:
            ttmp = teams[[team_id_col_teams, team_name_col_teams]].copy()
            ttmp[team_id_col_teams] = ttmp[team_id_col_teams].astype(str).str.strip()
            ttmp[team_name_col_teams] = ttmp[team_name_col_teams].astype(str).str.strip()
            ttmp = ttmp[(ttmp[team_id_col_teams] != "") & (ttmp[team_name_col_teams] != "")].drop_duplicates()
            team_id_to_name = dict(zip(ttmp[team_id_col_teams], ttmp[team_name_col_teams]))
            team_name_to_id = dict(zip(ttmp[team_name_col_teams], ttmp[team_id_col_teams]))

    if team_id_col_league and team_id_col_league in league.columns and team_id_to_name:
        league[team_id_col_league] = league[team_id_col_league].astype(str).str.strip()
        league["Team"] = league[team_id_col_league].map(team_id_to_name)
    elif "Team" not in league.columns:
        league["Team"] = None

    numeric_cols = [
        "Runs Scored",
        "Balls Faced",
        "6s",
        "Retirements",
        "Batting Strike Rate",
        "Batting Average",
        "Highest Score",
        "Innings Played",
        "Not Out's",
        "Total Overs",
        "Overs",
        "Balls Bowled",
        "Maidens",
        "Runs Conceded",
        "Wickets",
        "Wides",
        "No Balls",
        "Economy",
        "Bowling Strike Rate",
        "Bowling Average",
        "Catches",
        "Run Outs",
        "Stumpings",
        "Fantasy Points",
    ]
    for col in numeric_cols:
        if col in league.columns:
            league[col] = pd.to_numeric(league[col], errors="coerce")

    selected_team_id = None
    if enable_team_filter:
        team_names = sorted([t for t in team_name_to_id.keys() if str(t).strip() != ""]) if team_name_to_id else []
        team_dropdown_options = ["All"] + team_names
        c1, c2 = st.columns([2, 1])
        with c2:
            selected_team_name = st.selectbox(
                "Team",
                team_dropdown_options if team_dropdown_options else ["All"],
                key="ps_team_name",
            )
        selected_team_id = team_name_to_id.get(selected_team_name) if selected_team_name != "All" else None
    else:
        c1 = st.container()

    player_options_df = league
    if selected_team_id is not None and team_id_col_league and team_id_col_league in league.columns:
        player_options_df = league[
            league[team_id_col_league].astype(str).str.strip() == str(selected_team_id).strip()
        ]

    player_options = (
        player_options_df[name_col].dropna().astype(str).map(str.strip)
        if name_col and name_col in player_options_df.columns
        else pd.Series(dtype=str)
    )
    player_options_list = sorted([p for p in player_options.unique().tolist() if p != ""])

    current_players = st.session_state.get("ps_players", [])
    current_players = [p for p in current_players if p in player_options_list]
    st.session_state["ps_players"] = current_players

    with c1:
        selected_players = st.multiselect(
            "Players - Leave blank for all players",
            player_options_list,
            key="ps_players",
        )

    filtered = league.copy()
    if selected_team_id is not None and team_id_col_league and team_id_col_league in filtered.columns:
        filtered = filtered[filtered[team_id_col_league].astype(str).str.strip() == str(selected_team_id).strip()]
    if name_col and name_col in filtered.columns and selected_players:
        filtered = filtered[filtered[name_col].astype(str).str.strip().isin(selected_players)]

    BATTING_STATS = [
        "Runs Scored",
        "Balls Faced",
        "6s",
        "Retirements",
        "Batting Strike Rate",
        "Batting Average",
        "Highest Score",
        "Innings Played",
        "Not Out's",
    ]
    BOWLING_STATS = [
        "Overs",
        "Balls Bowled",
        "Maidens",
        "Runs Conceded",
        "Wickets",
        "Wides",
        "No Balls",
        "Economy",
        "Bowling Strike Rate",
        "Bowling Average",
        "Best Figures",
    ]
    FIELDING_STATS = ["Catches", "Run Outs", "Stumpings"]

    batting_options = [c for c in BATTING_STATS if c in filtered.columns]
    bowling_options = [c for c in BOWLING_STATS if c in filtered.columns]
    fielding_options = [c for c in FIELDING_STATS if c in filtered.columns]
    other_aliases = {
        "Fantasy Points": ["Fantasy Points", "Total Fantasy Points", "Fantasy Points Total", "Points"],
        "Average Fantasy Points": [
            "Average Fantasy Points",
            "Avg Fantasy Points",
            "Ave Fantasy Points",
            "Ave Points Per Match",
            "Avg Points Per Match",
        ],
        "Matches Played": ["Matches Played", "Match Played", "Games Played", "Played"],
    }
    other_display_to_actual: dict[str, str] = {}
    for display_name, aliases in other_aliases.items():
        mapped_col = _find_col(filtered, aliases)
        if mapped_col and mapped_col in filtered.columns:
            other_display_to_actual[display_name] = mapped_col
    other_options = [d for d in ["Fantasy Points", "Average Fantasy Points", "Matches Played"] if d in other_display_to_actual]

    st.markdown("#### Select Stats To Display")
    d1, d2, d3, d4 = st.columns(4)

    DEFAULT_BATTING = ["Runs Scored", "Batting Average"]
    DEFAULT_BOWLING = ["Wickets", "Economy"]
    DEFAULT_FIELDING: list[str] = []
    DEFAULT_OTHER: list[str] = []

    batting_select_key = "ps_batting_cols"
    bowling_select_key = "ps_bowling_cols"
    fielding_select_key = "ps_fielding_cols"
    other_select_key = "ps_other_cols"

    if f"__prev_{batting_select_key}" not in st.session_state:
        st.session_state[f"__prev_{batting_select_key}"] = DEFAULT_BATTING
    if f"__prev_{bowling_select_key}" not in st.session_state:
        st.session_state[f"__prev_{bowling_select_key}"] = DEFAULT_BOWLING
    if f"__prev_{fielding_select_key}" not in st.session_state:
        st.session_state[f"__prev_{fielding_select_key}"] = DEFAULT_FIELDING
    if f"__prev_{other_select_key}" not in st.session_state:
        st.session_state[f"__prev_{other_select_key}"] = DEFAULT_OTHER

    _init_or_sanitize_multiselect_state_allow_empty(batting_select_key, [ALL] + batting_options, DEFAULT_BATTING)
    _init_or_sanitize_multiselect_state_allow_empty(bowling_select_key, [ALL] + bowling_options, DEFAULT_BOWLING)
    _init_or_sanitize_multiselect_state_allow_empty(fielding_select_key, [ALL] + fielding_options, DEFAULT_FIELDING)
    _init_or_sanitize_multiselect_state_allow_empty(other_select_key, [ALL] + other_options, DEFAULT_OTHER)

    with d1:
        st.multiselect(
            "Batting stats",
            options=[ALL] + batting_options,
            key=batting_select_key,
            on_change=lambda: enforce_all_exclusive(batting_select_key),
        )
    with d2:
        st.multiselect(
            "Bowling stats",
            options=[ALL] + bowling_options,
            key=bowling_select_key,
            on_change=lambda: enforce_all_exclusive(bowling_select_key),
        )
    with d3:
        st.multiselect(
            "Fielding stats",
            options=[ALL] + fielding_options,
            key=fielding_select_key,
            on_change=lambda: enforce_all_exclusive(fielding_select_key),
        )
    with d4:
        st.multiselect(
            "Other",
            options=[ALL] + other_options,
            key=other_select_key,
            on_change=lambda: enforce_all_exclusive(other_select_key),
        )

    resolved_batting = _resolved_selection(batting_select_key, batting_options, DEFAULT_BATTING)
    resolved_bowling = _resolved_selection(bowling_select_key, bowling_options, DEFAULT_BOWLING)
    resolved_fielding = _resolved_selection(fielding_select_key, fielding_options, DEFAULT_FIELDING)
    resolved_other_display = _resolved_selection(other_select_key, other_options, DEFAULT_OTHER)
    resolved_other = [other_display_to_actual[d] for d in resolved_other_display if d in other_display_to_actual]

    st.session_state[f"__prev_{batting_select_key}"] = st.session_state.get(batting_select_key) or []
    st.session_state[f"__prev_{bowling_select_key}"] = st.session_state.get(bowling_select_key) or []
    st.session_state[f"__prev_{fielding_select_key}"] = st.session_state.get(fielding_select_key) or []
    st.session_state[f"__prev_{other_select_key}"] = st.session_state.get(other_select_key) or []

    selected_columns = resolved_batting + resolved_bowling + resolved_fielding + resolved_other

    fixed_cols: list[str] = []
    if "Name" in filtered.columns:
        fixed_cols.append("Name")
    elif name_col and name_col in filtered.columns:
        fixed_cols.append(name_col)
    display_cols: list[str] = []
    for c in fixed_cols:
        if c and c in filtered.columns and c not in display_cols:
            display_cols.append(c)
    for c in selected_columns:
        if c in filtered.columns and c not in display_cols:
            display_cols.append(c)

    view = filtered[display_cols].copy() if all(c in filtered.columns for c in display_cols) else filtered.copy()
    if "Fantasy Points" in view.columns:
        try:
            view = view.sort_values(by="Fantasy Points", ascending=False)
        except Exception:
            pass

    col_config: dict = {}
    if "Name" in view.columns:
        col_config["Name"] = st.column_config.TextColumn(pinned=True)
    elif name_col and name_col in view.columns:
        col_config[name_col] = st.column_config.TextColumn(pinned=True)
    for c in ["Batting Strike Rate", "Batting Average", "Economy", "Bowling Strike Rate", "Bowling Average"]:
        if c in view.columns:
            col_config[c] = st.column_config.NumberColumn(format="%.2f")
    avg_fantasy_ppm_cols = [
        "Average Fantasy Points per Match",
        "Average Fantasy Points",
        "Avg Fantasy Points",
        "Ave Fantasy Points",
        "Average Points Per Match",
        "Avg Points Per Match",
        "Ave Points Per Match",
    ]
    for c in avg_fantasy_ppm_cols:
        if c in view.columns:
            col_config[c] = st.column_config.NumberColumn(format="%.2f")
    if "Fantasy Points" in view.columns:
        col_config["Fantasy Points"] = st.column_config.NumberColumn()

    st.data_editor(
        view,
        width="stretch",
        hide_index=True,
        disabled=True,
        column_config=col_config,
    )


# ---- Read secrets ----
try:
    app_key = _get_secret("DROPBOX_APP_KEY")
    app_secret = _get_secret("DROPBOX_APP_SECRET")
    refresh_token = _get_secret("DROPBOX_REFRESH_TOKEN")
    dropbox_path = _get_secret("DROPBOX_FILE_PATH")
except Exception as e:
    st.error(str(e))
    st.stop()

# ---- Load workbook from Dropbox ----
with st.spinner("Loading latest league workbook from Dropbox..."):
    try:
        workbook_bytes = _download_workbook_bytes(app_key, app_secret, refresh_token, dropbox_path)
        data = load_league_workbook_from_bytes(workbook_bytes)
    except Exception as e:
        st.error(f"Failed to load workbook from Dropbox: {e}")
        st.stop()

# ---- Fixtures ----
fixtures = data.fixture_results.copy()
fixtures.columns = [str(c).strip() for c in fixtures.columns]

# ---- League table (pre-calculated in Excel) ----
league_table_df = getattr(data, "league_table", None)
if league_table_df is not None and not league_table_df.empty:
    league_table = league_table_df.copy()
    league_table.columns = [str(c).strip() for c in league_table.columns]
else:
    league_table = pd.DataFrame()

# ----------------------------
# Tabs
# ----------------------------
selected_tab = st.radio(
    label="Navigation Tabs",
    options=[
        "Fixtures & Results",
        "League Table",
        "Teams",
        "Player Stats",
        "Historical Stats",
        "Top Performers",
        "Scorecards",
    ],
    horizontal=True,
    key="main_tab",
    label_visibility="collapsed",
)

st.markdown(
    """
    <style>
    /* =========================================================
       Stateful tabs built from st.radio (native Streamlit look)
       ========================================================= */

    /* --- Radiogroup container (acts like tab bar) --- */
    div[role="radiogroup"] {
        display: flex !important;
        flex-direction: row !important;
        gap: 1.25rem !important;
        border-bottom: none !important;      /* no separator line */
        padding-bottom: 0 !important;
        margin-bottom: 1.25rem;
    }

    /* --- Each tab label --- */
    div[role="radiogroup"] > label {
        display: inline-flex !important;
        align-items: center !important;
        margin: 0 !important;
        padding: 0.45rem 0 !important;
        cursor: pointer !important;
        background: transparent !important;
        border: none !important;
        gap: 0 !important;

        /* Underline support (prevents layout quirks) */
        border-bottom: 2px solid transparent !important;
        text-decoration: none !important;
    }

    /* --- Hide radio controls ONLY (keep labels visible) --- */
    div[role="radiogroup"] > label > div:first-child,
    div[role="radiogroup"] > label > span:first-child {
        display: none !important;
        width: 0 !important;
        height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }

    div[role="radiogroup"] > label svg {
        display: none !important;
        width: 0 !important;
        height: 0 !important;
    }

    div[role="radiogroup"] input[type="radio"] {
        position: absolute !important;
        opacity: 0 !important;
        width: 0 !important;
        height: 0 !important;
        pointer-events: none !important;
    }

    /* --- Tab text container (Streamlit varies between div/span) --- */
    div[role="radiogroup"] > label > div,
    div[role="radiogroup"] > label > span {
        padding: 0 !important;
        font-weight: 500 !important;
        color: rgba(49, 51, 63, 0.75) !important;   /* unselected (light) */
    }

    /* Hover (light mode) */
    div[role="radiogroup"] > label:hover > div,
    div[role="radiogroup"] > label:hover > span {
        color: rgba(49, 51, 63, 1) !important;
    }

    /* Selected tab (light mode): underline + red text */
    div[role="radiogroup"] > label:has(input:checked) {
        border-bottom-color: rgba(255, 0, 0, 0.85) !important;
    }

    div[role="radiogroup"] > label:has(input:checked) > div,
    div[role="radiogroup"] > label:has(input:checked) > span {
        font-weight: 600 !important;
        color: rgba(255, 0, 0, 0.85) !important;    /* selected red */
    }

    /* =========================================================
       Dark mode: unselected white, selected red, underline red
       ========================================================= */
    @media (prefers-color-scheme: dark) {

        /* Unselected */
        div[role="radiogroup"] > label > div,
        div[role="radiogroup"] > label > span {
            color: rgba(255, 255, 255, 0.90) !important;
        }

        /* Hover */
        div[role="radiogroup"] > label:hover > div,
        div[role="radiogroup"] > label:hover > span {
            color: rgba(255, 255, 255, 1) !important;
        }

        /* Selected */
        div[role="radiogroup"] > label:has(input:checked) {
            border-bottom-color: rgba(255, 0, 0, 0.90) !important;
        }

        div[role="radiogroup"] > label:has(input:checked) > div,
        div[role="radiogroup"] > label:has(input:checked) > span {
            color: rgba(255, 0, 0, 0.90) !important;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================
# TAB: TOP PERFORMERS
# ============================
if selected_tab == "Top Performers":
    st.subheader("Top Performers")

    df_raw = getattr(data, "top_performers", None)
    if df_raw is None or df_raw.empty:
        st.info("Top performers are not available yet.")
        st.stop()

    parsed = parse_top_performers_grid(df_raw)
    render_top_performers_streamlit(parsed)

# ============================
# TAB 2: FIXTURES & RESULTS
# ============================
if selected_tab == "Fixtures & Results":
    st.subheader("Fixtures & Results")

    display = fixtures.copy()

    if "Date" in display.columns:
        display["Date"] = _format_date_dd_mmm(display["Date"])
    if "Time" in display.columns:
        display["Time"] = _format_time_ampm(display["Time"])

    ordered_cols = ["Date", "Time", "Home Team", "Away Team", "Status", "Won By", "Home Score", "Away Score"]
    show_cols = [c for c in ordered_cols if c in display.columns]

    st.dataframe(
        display[show_cols] if show_cols else display,
        width="stretch",
        hide_index=True,
    )

# ============================
# TAB 3: LEAGUE TABLE
# ============================
if selected_tab == "League Table":
    st.subheader("League Table")

    if league_table is None or league_table.empty:
        st.info(
            "League table not available yet. Confirm the Excel table is named 'League_Table' "
            "on sheet 'Fixture_Results' and that it contains at least one data row."
        )
    else:
        lt = league_table.copy()

        cols_to_hide = [
            "Runs Scored",
            "Runs Conceeded",
            "Wickets Taken",
            "Wickets Lost",
            "Overs Faced",
            "Overs Bowled",
        ]
        lt = lt.drop(columns=[c for c in cols_to_hide if c in lt.columns], errors="ignore")

        lt.insert(0, "Position", range(1, len(lt) + 1))

        if "NRR" in lt.columns:
            nrr = pd.to_numeric(lt["NRR"], errors="coerce")
            lt["NRR"] = nrr.map(lambda x: f"{x:.2f}" if pd.notna(x) else "")

        html_table = lt.to_html(index=False, escape=True)

        st.markdown(
            """
            <style>
                .lt-wrap {
                width: 100%;
                border: 1px solid rgba(49, 51, 63, 0.15);
                border-radius: 0.5rem;
                overflow: hidden;
                background: white;
                padding-bottom: 0;
              }

                .lt-scroll {
                width: 100%;
                overflow-x: auto;
                overflow-y: hidden;
              }

                .lt-wrap table {
                width: 100%;
                border-collapse: separate;
                border-spacing: 0;
                font-size: 0.95rem;
                border-top: none !important;
                border-bottom: none !important;
                margin: 0 !important;
                padding: 0 !important;
              }

                .lt-wrap thead th {
                position: sticky;
                top: 0;
                z-index: 2;
                background: rgba(250, 250, 252, 1);
                color: rgba(49, 51, 63, 0.9);
                text-align: left;
                font-weight: 600;
                padding: 0.65rem 0.75rem;
                border-bottom: 1px solid rgba(49, 51, 63, 0.15);
                white-space: nowrap;
              }

                .lt-wrap tbody td {
                padding: 0.6rem 0.75rem;
                border-bottom: 1px solid rgba(49, 51, 63, 0.08);
                color: rgba(49, 51, 63, 0.95);
                white-space: nowrap;
              }

                .lt-wrap tbody td:not(:nth-child(2)),
                .lt-wrap thead th:not(:nth-child(2)) {
                text-align: center;
              }

                .lt-wrap tbody tr:nth-child(1) td { background: rgba(255, 215, 0, 0.08); }
                .lt-wrap tbody tr:nth-child(2) td { background: rgba(192, 192, 192, 0.22); }
                .lt-wrap tbody tr:nth-child(3) td { background: rgba(205, 127, 50, 0.10); }

                .lt-wrap tbody tr:last-child td { border-bottom: 1px solid transparent; }

                .lt-wrap tbody tr:hover td { background: rgba(240, 242, 246, 1); }

                .lt-wrap table, .lt-wrap th, .lt-wrap td {
                border-left: none !important;
                border-right: none !important;
              }

            @media (prefers-color-scheme: dark) {

                .lt-wrap {
                    background: rgba(14, 17, 23, 1) !important;
                    border: 1px solid rgba(255, 255, 255, 0.12) !important;
                }

                .lt-wrap table {
                    background: transparent !important;
                }

                .lt-wrap thead th {
                    background: rgba(28, 31, 38, 1) !important;
                    color: rgba(255, 255, 255, 0.90) !important;
                    border-bottom: 1px solid rgba(255, 255, 255, 0.12) !important;
                }

                .lt-wrap tbody td {
                    color: rgba(255, 255, 255, 0.88) !important;
                    border-bottom: 1px solid rgba(255, 255, 255, 0.08) !important;
                }

                .lt-wrap tbody tr:hover td {
                    background: rgba(255, 255, 255, 0.06) !important;
                }

                .lt-wrap tbody tr:nth-child(1) td { background: rgba(255, 215, 0, 0.10) !important; }
                .lt-wrap tbody tr:nth-child(2) td { background: rgba(192, 192, 192, 0.10) !important; }
                .lt-wrap tbody tr:nth-child(3) td { background: rgba(205, 127, 50, 0.10) !important; }

                .lt-wrap table, .lt-wrap th, .lt-wrap td {
                    border-left: none !important;
                    border-right: none !important;
                }
            }

            </style>
            """,
            unsafe_allow_html=True,
        )

        st.markdown(
            f"""
            <div class="lt-wrap">
              <div class="lt-scroll">
                {html_table}
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )


# ============================
# TAB 3: TEAMS
# ============================
if selected_tab == "Teams":
    st.subheader("Teams")

    teams_df = getattr(data, "teams", None)
    league_df = getattr(data, "league_data", None)

    if teams_df is None or teams_df.empty:
        st.info("No Teams_Table found yet.")
        st.stop()

    teams = teams_df.copy()
    teams.columns = [str(c).strip() for c in teams.columns]

    team_id_col = _find_col(teams, ["TeamID", "Team Id", "Team ID"])
    team_name_col = _find_col(teams, ["Team Names", "Team Name"])
    active_col = _find_col(teams, ["Active"])
    captain_name_col = _find_col(teams, ["Captain's Name", "Captains Name", "Captain Name"])

    if not team_name_col:
        st.error("Teams_Table is missing 'Team Names'.")
        st.stop()

    teams[team_name_col] = teams[team_name_col].astype(str).str.strip()
    if team_id_col and team_id_col in teams.columns:
        teams[team_id_col] = teams[team_id_col].astype(str).str.strip()
    if active_col and active_col in teams.columns:
        teams[active_col] = teams[active_col].astype(str).str.strip()
    if captain_name_col and captain_name_col in teams.columns:
        teams[captain_name_col] = teams[captain_name_col].astype(str).str.strip()

    team_names = sorted(
        [t for t in teams[team_name_col].dropna().unique().tolist() if str(t).strip() != ""],
        key=str.lower,
    )
    team_choice = st.selectbox("Team", ["All Teams"] + team_names, key="ts_team_name")

    # ---------------------------------------------------------
    # ALL TEAMS
    # ---------------------------------------------------------
    if team_choice == "All Teams":

        if league_df is None or league_df.empty:
            st.info("No League_Data_Stats found yet, so team totals cannot be calculated.")
            st.stop()

        league = league_df.copy()
        league.columns = [str(c).strip() for c in league.columns]

        team_id_col_league = _find_col(league, ["TeamID", "Team Id", "Team ID"])

        team_id_to_name: dict[str, str] = {}
        if team_id_col and team_id_col in teams.columns:
            tmap = teams[[team_id_col, team_name_col]].copy()
            tmap[team_id_col] = tmap[team_id_col].astype(str).str.strip()
            tmap[team_name_col] = tmap[team_name_col].astype(str).str.strip()
            tmap = tmap[(tmap[team_id_col] != "") & (tmap[team_name_col] != "")].drop_duplicates()
            team_id_to_name = dict(zip(tmap[team_id_col], tmap[team_name_col]))

        if not team_id_col_league or team_id_col_league not in league.columns or not team_id_to_name:
            st.info("Team totals require TeamID in League_Data and TeamID/Team Names in Teams_Table.")
            st.stop()

        league[team_id_col_league] = league[team_id_col_league].astype(str).str.strip()
        league["Team"] = league[team_id_col_league].map(team_id_to_name)

        league = league[league["Team"].notna() & (league["Team"].astype(str).str.strip() != "")]
        if league.empty:
            st.info("No mapped team stats available yet.")
            st.stop()

        sum_cols = [
            "Runs Scored",
            "Balls Faced",
            "6s",
            "Retirements",
            "Innings Played",
            "Not Out's",
            "Total Overs",
            "Overs",
            "Balls Bowled",
            "Maidens",
            "Runs Conceded",
            "Wickets",
            "Wides",
            "No Balls",
            "Catches",
            "Run Outs",
            "Stumpings",
            "Fantasy Points",
        ]
        for c in sum_cols:
            if c in league.columns:
                league[c] = pd.to_numeric(league[c], errors="coerce")

        agg_map = {c: "sum" for c in sum_cols if c in league.columns}
        team_totals = league.groupby("Team", as_index=False).agg(agg_map) if agg_map else league[["Team"]].drop_duplicates()

        # Derived metrics (same column names as player stats where possible)
        if "Runs Scored" in team_totals.columns and "Balls Faced" in team_totals.columns:
            rs = pd.to_numeric(team_totals["Runs Scored"], errors="coerce")
            bf = pd.to_numeric(team_totals["Balls Faced"], errors="coerce")
            team_totals["Batting Strike Rate"] = (rs / bf) * 100
            team_totals.loc[(bf.isna()) | (bf <= 0), "Batting Strike Rate"] = pd.NA

        if "Runs Scored" in team_totals.columns and "Innings Played" in team_totals.columns and "Not Out's" in team_totals.columns:
            rs = pd.to_numeric(team_totals["Runs Scored"], errors="coerce")
            inn = pd.to_numeric(team_totals["Innings Played"], errors="coerce")
            no = pd.to_numeric(team_totals["Not Out's"], errors="coerce")
            outs = inn - no
            outs = outs.where((outs.notna()) & (outs > 0), 1)
            team_totals["Batting Average"] = rs / outs
            team_totals.loc[rs.isna(), "Batting Average"] = pd.NA

        if "Runs Conceded" in team_totals.columns and "Overs" in team_totals.columns:
            rc = pd.to_numeric(team_totals["Runs Conceded"], errors="coerce")
            ov = pd.to_numeric(team_totals["Overs"], errors="coerce")
            team_totals["Economy"] = rc / ov
            team_totals.loc[(ov.isna()) | (ov <= 0), "Economy"] = pd.NA

        if "Balls Bowled" in team_totals.columns and "Wickets" in team_totals.columns:
            bb = pd.to_numeric(team_totals["Balls Bowled"], errors="coerce")
            wk = pd.to_numeric(team_totals["Wickets"], errors="coerce")
            team_totals["Bowling Strike Rate"] = bb / wk
            team_totals.loc[(wk.isna()) | (wk <= 0), "Bowling Strike Rate"] = pd.NA

        if "Runs Conceded" in team_totals.columns and "Wickets" in team_totals.columns:
            rc = pd.to_numeric(team_totals["Runs Conceded"], errors="coerce")
            wk = pd.to_numeric(team_totals["Wickets"], errors="coerce")
            team_totals["Bowling Average"] = rc / wk
            team_totals.loc[(wk.isna()) | (wk <= 0), "Bowling Average"] = pd.NA

        # Join Active + Captain (optional)
        teams_named = teams.rename(columns={team_name_col: "Team"}).copy()
        teams_named["Team"] = teams_named["Team"].astype(str).str.strip()

        meta_cols: list[str] = []
        if active_col and active_col in teams_named.columns:
            meta_cols.append(active_col)

        tmeta = teams_named[["Team"] + meta_cols].drop_duplicates() if meta_cols else teams_named[["Team"]].drop_duplicates()
        team_totals = team_totals.merge(tmeta, on="Team", how="left")

        # ---- Form (Last 5) from Fixture_Results_Table ----
        # Uses: fixtures (already loaded at top), columns: Date, Time, Home Team, Away Team, Status, Won By
        def _team_form_last_n(team_name: str, n: int = 5) -> str:
            if fixtures is None or fixtures.empty:
                return ""

            f = fixtures.copy()
            f.columns = [str(c).strip() for c in f.columns]

            required = ["Date", "Time", "Home Team", "Away Team", "Status", "Won By"]
            if not all(c in f.columns for c in required):
                return ""

            # Only matches involving this team
            f = f[
                (f["Home Team"].astype(str).str.strip() == str(team_name).strip())
                | (f["Away Team"].astype(str).str.strip() == str(team_name).strip())
            ].copy()

            if f.empty:
                return ""

            # Sort by Date+Time (most recent first)
            dt = pd.to_datetime(f["Date"], errors="coerce")
            tm = pd.to_datetime("2000-01-01 " + f["Time"].astype(str), errors="coerce").dt.time
            f["_dt"] = pd.to_datetime(dt.dt.date.astype(str) + " " + tm.astype(str), errors="coerce")
            f = f.sort_values("_dt", ascending=False)

            # Keep only completed matches
            f = f[f["Status"].astype(str).str.strip().isin(["Played", "Abandoned"])].copy()

            # If no completed matches, show nothing
            if f.empty:
                return ""

            # Take up to last N completed matches (most recent N)
            f = f.head(n)

            out = []
            team_s = str(team_name).strip().lower()

            for _, r in f.iterrows():
                status = str(r.get("Status", "")).strip()
                won_by = str(r.get("Won By", "")).strip().lower()

                # Abandoned is always a dash in the form guide
                if status == "Abandoned":
                    out.append("➖")
                    continue

                # Played: decide W/L if possible, else dash
                if won_by and team_s in won_by:
                    out.append("✅")
                elif won_by:
                    out.append("❌")
                else:
                    out.append("➖")

            return " ".join(out)

        # Compute form per team
        team_totals["Form (Last 5)"] = team_totals["Team"].apply(lambda t: _team_form_last_n(t, 5))

        # ---- Sort All Teams to match league_table order (as sorted in Excel) ----
        # Requires league_table to contain a Team column with the same labels as team_totals["Team"]
        if "Team" in league_table.columns and not league_table.empty:
            _order_df = league_table[["Team"]].copy()
            _order_df["__order"] = range(len(_order_df))
            team_totals = team_totals.merge(_order_df, on="Team", how="left")
            team_totals = team_totals.sort_values("__order", ascending=True, na_position="last").drop(columns=["__order"])

        # ---- selectors (Batting / Bowling / Fielding) ----
        TEAM_BATTING_STATS = [
            "Runs Scored",
            "Balls Faced",
            "6s",
            "Retirements",
            "Batting Strike Rate",
            "Batting Average",
        ]
        TEAM_BOWLING_STATS = [
            "Overs",
            "Balls Bowled",
            "Maidens",
            "Runs Conceded",
            "Wickets",
            "Wides",
            "No Balls",
            "Economy",
            "Bowling Strike Rate",
            "Bowling Average",
        ]
        TEAM_FIELDING_STATS = ["Catches", "Run Outs", "Stumpings"]

        batting_options = [c for c in TEAM_BATTING_STATS if c in team_totals.columns]
        bowling_options = [c for c in TEAM_BOWLING_STATS if c in team_totals.columns]
        fielding_options = [c for c in TEAM_FIELDING_STATS if c in team_totals.columns]

        default_batting = [c for c in ["Runs Scored", "Batting Average"] if c in batting_options]
        default_bowling = [c for c in ["Wickets", "Economy"] if c in bowling_options]
        default_fielding: list[str] = []

        _init_or_sanitize_multiselect_state_allow_empty("ts_all_batting_cols", batting_options, default_batting)
        _init_or_sanitize_multiselect_state_allow_empty("ts_all_bowling_cols", bowling_options, default_bowling)
        _init_or_sanitize_multiselect_state_allow_empty("ts_all_fielding_cols", fielding_options, default_fielding)

        st.markdown("#### Select Team Stats To Display")
        d1, d2, d3 = st.columns(3)
        with d1:
            selected_batting = st.multiselect("Batting Stats", options=batting_options, key="ts_all_batting_cols")
        with d2:
            selected_bowling = st.multiselect("Bowling Stats", options=bowling_options, key="ts_all_bowling_cols")
        with d3:
            selected_fielding = st.multiselect("Fielding Stats", options=fielding_options, key="ts_all_fielding_cols")

        selected_columns = selected_batting + selected_bowling + selected_fielding

        # Build columns: Team + Form + meta + selected + Fantasy Points (Fantasy Points last)
        display_cols = ["Team"]
        if "Form (Last 5)" in team_totals.columns:
            display_cols.append("Form (Last 5)")

        for mc in meta_cols:
            if mc in team_totals.columns and mc not in display_cols:
                display_cols.append(mc)

        for c in selected_columns:
            if c in team_totals.columns and c not in display_cols:
                display_cols.append(c)

        if "Fantasy Points" in team_totals.columns and "Fantasy Points" not in display_cols:
            display_cols.append("Fantasy Points")

        view = team_totals[display_cols].copy() if all(c in team_totals.columns for c in display_cols) else team_totals.copy()

        col_config = {"Team": st.column_config.TextColumn(pinned=True)}
        for c in ["Batting Strike Rate", "Batting Average", "Economy", "Bowling Strike Rate", "Bowling Average"]:
            if c in view.columns:
                col_config[c] = st.column_config.NumberColumn(format="%.2f")

        # Do not pin Fantasy Points (ensures it stays far right)
        if "Fantasy Points" in view.columns:
            col_config["Fantasy Points"] = st.column_config.NumberColumn()

        st.data_editor(
            view,
            width="stretch",
            hide_index=True,
            disabled=True,
            column_config=col_config,
        )

        st.markdown("---")
        st.caption("Select a team above to view team details.")
        st.stop()
    # ---------------------------------------------------------
    # SINGLE TEAM VIEW
    # ---------------------------------------------------------
    team_row = teams.loc[teams[team_name_col] == team_choice]
    if team_row.empty:
        st.info("Selected team not found in Teams_Table.")
        st.stop()
    team_row = team_row.iloc[0]

    meta_c1, meta_c2, meta_c3 = st.columns([2, 1, 2])
    with meta_c1:
        st.markdown(f"**Team:** {team_choice}")
    with meta_c2:
        st.markdown(f"**Active:** {team_row.get(active_col, '—') if active_col else '—'}")
    with meta_c3:
        st.markdown(f"**Captain:** {team_row.get(captain_name_col, '—') if captain_name_col else '—'}")

    if league_table is not None and not league_table.empty and "Team" in league_table.columns:
        lt_lookup = league_table.copy()
        lt_lookup.columns = [str(c).strip() for c in lt_lookup.columns]
        lt_team = lt_lookup[lt_lookup["Team"].astype(str).str.strip() == str(team_choice).strip()]
        if not lt_team.empty:
            r = lt_team.iloc[0].to_dict()
            played = r.get("Played", "—")
            points = r.get("Points", "—")
            nrr_val = r.get("NRR", "—")
            try:
                nrr_val = f"{float(nrr_val):.2f}"
            except Exception:
                pass
            st.markdown(f"**Played:** {played} &nbsp;&nbsp; **Points:** {points} &nbsp;&nbsp; **NRR:** {nrr_val}")

    st.markdown("---")

    if league_df is None or league_df.empty:
        st.info("No League_Data_Stats found yet, so team stats cannot be displayed.")
        st.stop()

    league = league_df.copy()
    league.columns = [str(c).strip() for c in league.columns]

    name_col = _find_col(league, ["Name"])
    team_id_col_league = _find_col(league, ["TeamID", "Team Id", "Team ID"])

    if not (team_id_col and team_id_col_league and team_id_col in teams.columns and team_id_col_league in league.columns):
        st.info("Team page requires TeamID in Teams_Table and League_Data.")
        st.stop()

    selected_team_id = str(team_row.get(team_id_col, "")).strip()
    if not selected_team_id:
        st.info("Selected team has no TeamID in Teams_Table.")
        st.stop()

    filtered_team = league.copy()
    filtered_team[team_id_col_league] = filtered_team[team_id_col_league].astype(str).str.strip()
    filtered_team = filtered_team[filtered_team[team_id_col_league] == selected_team_id]

    if filtered_team.empty:
        st.info("No matching player stats found for this team yet.")
        st.stop()

    numeric_cols = [
        "Runs Scored", "Balls Faced", "6s", "Retirements",
        "Batting Strike Rate", "Batting Average", "Highest Score",
        "Innings Played", "Not Out's",
        "Total Overs", "Overs", "Balls Bowled", "Maidens", "Runs Conceded", "Wickets",
        "Wides", "No Balls", "Economy", "Bowling Strike Rate", "Bowling Average",
        "Catches", "Run Outs", "Stumpings", "Fantasy Points",
    ]
    for c in numeric_cols:
        if c in filtered_team.columns:
            filtered_team[c] = pd.to_numeric(filtered_team[c], errors="coerce")

    # Selectors (Batting / Bowling / Fielding)
    BATTING_STATS = [
        "Runs Scored",
        "Balls Faced",
        "6s",
        "Retirements",
        "Batting Strike Rate",
        "Batting Average",
        "Highest Score",
        "Innings Played",
        "Not Out's",
    ]
    BOWLING_STATS = [
        "Total Overs",
        "Overs",
        "Balls Bowled",
        "Maidens",
        "Runs Conceded",
        "Wickets",
        "Wides",
        "No Balls",
        "Economy",
        "Bowling Strike Rate",
        "Bowling Average",
        "Best Figures",
    ]
    FIELDING_STATS = ["Catches", "Run Outs", "Stumpings"]

    batting_options = [c for c in BATTING_STATS if c in filtered_team.columns]
    bowling_options = [c for c in BOWLING_STATS if c in filtered_team.columns]
    fielding_options = [c for c in FIELDING_STATS if c in filtered_team.columns]

    default_batting = [c for c in ["Runs Scored", "Batting Average"] if c in batting_options]
    default_bowling = [c for c in ["Wickets", "Economy"] if c in bowling_options]
    default_fielding: list[str] = []

    _init_or_sanitize_multiselect_state_allow_empty("ts_batting_cols", batting_options, default_batting)
    _init_or_sanitize_multiselect_state_allow_empty("ts_bowling_cols", bowling_options, default_bowling)
    _init_or_sanitize_multiselect_state_allow_empty("ts_fielding_cols", fielding_options, default_fielding)

    st.markdown("#### Select Stats To Display")
    d1, d2, d3 = st.columns(3)
    with d1:
        selected_batting = st.multiselect("Batting Stats", options=batting_options, key="ts_batting_cols")
    with d2:
        selected_bowling = st.multiselect("Bowling Stats", options=bowling_options, key="ts_bowling_cols")
    with d3:
        selected_fielding = st.multiselect("Fielding Stats", options=fielding_options, key="ts_fielding_cols")

    selected_columns = selected_batting + selected_bowling + selected_fielding

    fixed_name = "Name" if "Name" in filtered_team.columns else (name_col if name_col in filtered_team.columns else None)
    fixed_cols: list[str] = []
    if fixed_name:
        fixed_cols.append(fixed_name)

    display_cols: list[str] = []
    for c in fixed_cols:
        if c in filtered_team.columns and c not in display_cols:
            display_cols.append(c)

    for c in selected_columns:
        if c in filtered_team.columns and c not in display_cols:
            display_cols.append(c)

    if "Fantasy Points" in filtered_team.columns and "Fantasy Points" not in display_cols:
        display_cols.append("Fantasy Points")

    player_view = filtered_team[display_cols].copy() if display_cols else filtered_team.copy()

    if "Fantasy Points" in player_view.columns:
        try:
            player_view = player_view.sort_values(by="Fantasy Points", ascending=False)
        except Exception:
            pass
    elif "Runs Scored" in player_view.columns:
        try:
            player_view = player_view.sort_values(by="Runs Scored", ascending=False)
        except Exception:
            pass

    col_config: dict = {}
    if fixed_name and fixed_name in player_view.columns:
        col_config[fixed_name] = st.column_config.TextColumn(pinned=True)

    for c in ["Batting Strike Rate", "Batting Average", "Economy", "Bowling Strike Rate", "Bowling Average"]:
        if c in player_view.columns:
            col_config[c] = st.column_config.NumberColumn(format="%.2f")

    # Do not pin Fantasy Points (ensures it stays far right)
    if "Fantasy Points" in player_view.columns:
        col_config["Fantasy Points"] = st.column_config.NumberColumn()

    st.markdown("#### Player Stats (Team)")
    st.data_editor(
        player_view,
        width="stretch",
        hide_index=True,
        disabled=True,
        column_config=col_config,
    )

    # Totals table (same columns as player_view)
    st.markdown("#### Team Totals")

    base = filtered_team.copy()

    def _sum(col: str) -> float | None:
        if col not in base.columns:
            return None
        s = pd.to_numeric(base[col], errors="coerce").fillna(0).sum()
        return float(s)

    totals_row: dict = {}
    if fixed_name:
        totals_row[fixed_name] = "Team Totals"

    for col in [
        "Runs Scored", "Balls Faced", "6s", "Retirements",
        "Innings Played", "Not Out's",
        "Total Overs", "Overs", "Balls Bowled", "Maidens", "Runs Conceded", "Wickets", "Wides", "No Balls",
        "Catches", "Run Outs", "Stumpings", "Fantasy Points",
    ]:
        val = _sum(col)
        if val is not None and col in player_view.columns:
            totals_row[col] = val

    if "Batting Strike Rate" in player_view.columns:
        rs = _sum("Runs Scored") or 0.0
        bf = _sum("Balls Faced") or 0.0
        totals_row["Batting Strike Rate"] = (rs / bf) * 100 if bf > 0 else pd.NA

    if "Batting Average" in player_view.columns:
        rs = _sum("Runs Scored") or 0.0
        inn = _sum("Innings Played")
        no = _sum("Not Out's")
        if inn is not None and no is not None:
            outs = inn - no
            outs = outs if outs > 0 else 1.0
            totals_row["Batting Average"] = rs / outs
        else:
            totals_row["Batting Average"] = pd.NA

    if "Economy" in player_view.columns:
        rc = _sum("Runs Conceded") or 0.0
        ov = _sum("Overs") or 0.0
        totals_row["Economy"] = (rc / ov) if ov > 0 else pd.NA

    if "Bowling Strike Rate" in player_view.columns:
        bb = _sum("Balls Bowled") or 0.0
        wk = _sum("Wickets") or 0.0
        totals_row["Bowling Strike Rate"] = (bb / wk) if wk > 0 else pd.NA

    if "Bowling Average" in player_view.columns:
        rc = _sum("Runs Conceded") or 0.0
        wk = _sum("Wickets") or 0.0
        totals_row["Bowling Average"] = (rc / wk) if wk > 0 else pd.NA

    totals_df = pd.DataFrame([{c: totals_row.get(c, pd.NA) for c in player_view.columns}])

    st.data_editor(
        totals_df,
        width="stretch",
        hide_index=True,
        disabled=True,
        column_config=col_config,
    )
# ============================
# TAB 5/6: PLAYER & HISTORICAL STATS
# ============================
if selected_tab in ("Player Stats", "Historical Stats"):
    st.subheader(selected_tab)
    current_league_df = getattr(data, "league_data", None)
    current_teams_df = _extract_teams_df(data)
    season_map = discover_seasons()
    season_options = list(season_map.keys()) + ["All Stats"]
    selected_season = st.selectbox("Season", season_options, key="season_select")

    if selected_season == "Current Season":
        current_sheet_df = load_stats_for_sheet(workbook_bytes, season_map["Current Season"])
        if (current_sheet_df is None or current_sheet_df.empty) and current_league_df is not None and not current_league_df.empty:
            current_sheet_df = current_league_df.copy()
        if current_sheet_df is None or current_sheet_df.empty:
            st.warning("Sheet 'League_Data' is missing or has no player stats.")
        else:
            render_player_stats_ui(
                df=current_sheet_df,
                enable_team_filter=True,
                current_season=True,
                teams_df=current_teams_df,
                season_label=selected_season,
            )
    elif selected_season == "All Stats":
        try:
            all_stats_df = load_combined_stats_table(workbook_bytes, table_name="Combined_Stats")
        except Exception:
            st.error("Combined_Stats table not found in workbook")
            all_stats_df = pd.DataFrame()

        if all_stats_df is not None and not all_stats_df.empty:
            render_player_stats_ui(
                df=all_stats_df,
                enable_team_filter=False,
                current_season=False,
                teams_df=None,
                season_label=selected_season,
            )
        elif all_stats_df is not None and all_stats_df.empty:
            st.info("Combined_Stats table is empty or contains no valid players.")
    else:
        sheet_name = season_map.get(selected_season, "")
        season_df = load_stats_for_sheet(workbook_bytes, sheet_name) if sheet_name else pd.DataFrame()
        if season_df is None or season_df.empty:
            st.warning(f"Sheet '{sheet_name}' is missing or has no valid player stats.")
        else:
            render_player_stats_ui(
                df=season_df,
                enable_team_filter=False,
                current_season=False,
                teams_df=None,
                season_label=selected_season,
            )
# ============================
# TAB 5: SCORECARDS
# ============================
if selected_tab == "Scorecards":
    st.subheader("Scorecards")

    st.markdown("---")

    if "MatchID" not in fixtures.columns:
        st.info("Scorecards are not available because this workbook does not contain a 'MatchID' column.")
        st.stop()

    # Build a friendly fixture selector (Option A)
    fsel = fixtures.copy()
    fsel.columns = [str(c).strip() for c in fsel.columns]

    # Format date/time for display if present
    if "Date" in fsel.columns:
        fsel["Date"] = _format_date_dd_mmm(fsel["Date"])
    if "Time" in fsel.columns:
        fsel["Time"] = _format_time_ampm(fsel["Time"])

    def _safe(v) -> str:
        if pd.isna(v):
            return ""
        return str(v).strip()


    # Build all fixture options
    options: list[str] = []
    option_to_match: dict[str, str] = {}

    for _, r in fsel.iterrows():
        mid = _safe(r.get("MatchID"))
        if not mid:
            continue

        # Dropdown label for normal users: "01-Jan - 7 PM - Home vs Away"
        date_txt = _safe(r.get("Date")) if "Date" in fsel.columns else ""

        # Time is already formatted earlier via _format_time_ampm; reduce to "H AM/PM"
        time_txt = _safe(r.get("Time")) if "Time" in fsel.columns else ""
        if time_txt:
            # Examples handled: "7:00 PM" -> "7 PM", "7 PM" -> "7 PM"
            t = time_txt.replace(".", "").strip()
            if ":" in t:
                # "7:00 PM" -> ["7", "00 PM"] -> hour="7", ampm="PM"
                hour_part = t.split(":", 1)[0].strip()
                ampm_part = t.split(" ", 1)[-1].strip().upper()  # "PM"/"AM"
                time_txt = f"{hour_part} {ampm_part}"
            else:
                # Ensure spacing/case is consistent if already "7 PM"
                parts = t.split()
                if len(parts) >= 2:
                    time_txt = f"{parts[0]} {parts[-1].upper()}"

        match_txt = ""
        if "Home Team" in fsel.columns and "Away Team" in fsel.columns:
            match_txt = f"{_safe(r.get('Home Team'))} vs {_safe(r.get('Away Team'))}"

        label = " - ".join([p for p in [date_txt, time_txt, match_txt] if p])

        options.append(label)
        option_to_match[label] = mid

    if not options:
        st.info("No fixtures with a valid MatchID were found.")
        st.stop()

    # -------------------------------------------------
    # Fast filter: one DB query for all MatchIDs that have scorecards
    # -------------------------------------------------
    match_ids_with_scorecards = set(list_scorecard_match_ids())

    filtered_options = [label for label in options if option_to_match[label] in match_ids_with_scorecards]

    if not filtered_options:
        st.info("No scorecards have been uploaded for any fixtures yet.")
        st.stop()

    selected_fixture = st.selectbox(
        "Select a fixture to view available scorecards",
        filtered_options,
        key="fixtures_scorecard_select",
    )
    selected_match_id = option_to_match[selected_fixture]

    available = sorted(
        list_scorecards(selected_match_id),
        key=lambda r: (str(r.get("uploaded_at") or ""), int(r.get("scorecard_id") or 0)),
    )
    if not available:
        st.info("No scorecards have been uploaded for this fixture yet.")
        st.stop()

    st.caption(f"{len(available)} file(s) available")

    # -----------------------------
    # Image viewer (press & hold on mobile) - buttons only
    # -----------------------------
    image_rows = []
    for row in available:
        fname = (row.get("file_name") or "").strip()
        if fname.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
            image_rows.append(row)

    if image_rows:
        st.caption("Mobile (iPhone): press and hold the image to ‘Save to Photos’.")

        # Persistent index per match
        idx_key = f"scorecard_img_idx_{selected_match_id}"
        if idx_key not in st.session_state:
            st.session_state[idx_key] = 0

        # Clamp index (in case images were added/removed)
        n = len(image_rows)
        st.session_state[idx_key] = max(0, min(int(st.session_state[idx_key]), n - 1))
        idx = int(st.session_state[idx_key])

        # Render current image
        row = image_rows[idx]
        raw_fname = (row.get("file_name") or f"scorecard_{idx+1}").strip()

        # Remove file extension for display only
        fname = raw_fname.rsplit(".", 1)[0] if "." in raw_fname else raw_fname
        dbx_path = row.get("dropbox_path")

        if dbx_path:
            try:
                img_bytes = _download_scorecard_bytes(app_key, app_secret, refresh_token, dbx_path)
                st.markdown(f"**{fname}**")
                st.image(img_bytes, width="stretch")
            except Exception as e:
                st.warning(f"Could not load image '{fname}': {e}")

        # Image position indicator directly under the image
        st.caption(f"Image {idx + 1} of {n}")

        # Navigation buttons BELOW the caption
        c_prev, c_next = st.columns([1, 1])

        with c_prev:
            if st.button(
                "◀ Previous",
                width="stretch",
                disabled=(idx == 0),
                key=f"img_prev_{selected_match_id}",
            ):
                st.session_state[idx_key] = max(0, idx - 1)

        with c_next:
            if st.button(
                "Next ▶",
                width="stretch",
                disabled=(idx >= n - 1),
                key=f"img_next_{selected_match_id}",
            ):
                st.session_state[idx_key] = min(n - 1, idx + 1)

    # -----------------------------
    # PDFs (open in a new tab via Dropbox temporary link)
    # -----------------------------
    pdf_rows = []
    for row in available:
        fname = (row.get("file_name") or "").strip()
        if fname.lower().endswith(".pdf"):
            pdf_rows.append(row)

    if pdf_rows:
        st.markdown("#### PDFs")
        st.caption("Tap/click a PDF to open it in a new tab. Links are temporary.")

        for i, row in enumerate(pdf_rows):
            fname = (row.get("file_name") or f"scorecard_{i+1}.pdf").strip()
            dbx_path = row.get("dropbox_path")
            if not dbx_path:
                continue

            try:
                url = _get_temp_link(app_key, app_secret, refresh_token, dbx_path)
                # (1) Explicit key to prevent duplicate widget ID issues
                st.link_button(
                    f"{fname}",
                    url,
                    width="stretch",
                )
            except Exception as e:
                st.warning(f"Could not create link for '{fname}': {e}")
