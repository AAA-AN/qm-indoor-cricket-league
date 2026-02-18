from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Optional, List
import warnings

import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
)


@dataclass(frozen=True)
class ExcelLoadResult:
    fixture_results: pd.DataFrame
    league_table: Optional[pd.DataFrame] = None
    players: Optional[pd.DataFrame] = None
    teams: Optional[pd.DataFrame] = None
    league_data: Optional[pd.DataFrame] = None
    history_A_25_26: Optional[pd.DataFrame] = None
    history_B_24_25: Optional[pd.DataFrame] = None
    combined_stats: Optional[pd.DataFrame] = None
    top_performers: Optional[pd.DataFrame] = None


def _read_named_table(
    wb,
    sheet_name: str,
    table_name: str,
    *,
    drop_empty_columns: bool = True,
) -> pd.DataFrame:
    """
    Read an Excel Table (ListObject) by name and return a DataFrame.
    Uses the table ref so it stays correct even if the table range moves.

    drop_empty_columns:
      - True  -> drop columns that are entirely empty (all None/NaN)
      - False -> keep columns even if entirely empty (important for fixtures schema)
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = wb[sheet_name]
    if table_name not in ws.tables:
        available = ", ".join(sorted(ws.tables.keys())) if ws.tables else "(none)"
        raise ValueError(
            f"Table '{table_name}' not found on sheet '{sheet_name}'. Tables found: {available}"
        )

    table = ws.tables[table_name]
    ref = table.ref  # e.g. "B2:T32"
    cells = ws[ref]

    data: List[List[object]] = []
    for row in cells:
        data.append([c.value for c in row])

    if not data or len(data) < 2:
        raise ValueError(f"Table '{table_name}' appears to be empty.")

    headers = [str(h).strip() if h is not None else "" for h in data[0]]
    rows = data[1:]

    df = pd.DataFrame(rows, columns=headers)

    # Always drop columns with blank header names (these are never useful)
    blank_header_cols = [c for c in df.columns if str(c).strip() == ""]
    if blank_header_cols:
        df = df.drop(columns=blank_header_cols)

    # Only drop fully empty columns if requested
    if drop_empty_columns:
        df = df.dropna(axis=1, how="all")

    return df


def _read_named_table_any_sheet(
    wb,
    table_name: str,
    *,
    drop_empty_columns: bool = True,
) -> pd.DataFrame:
    """
    Read a named table by searching all sheets.
    Useful for history tables that may live on different sheets.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if table_name in ws.tables:
            return _read_named_table(
                wb,
                sheet_name=sheet_name,
                table_name=table_name,
                drop_empty_columns=drop_empty_columns,
            )
    raise ValueError(f"Table '{table_name}' not found in any worksheet.")


def _read_defined_name_range(wb, defined_name: str) -> pd.DataFrame:
    """
    Read an Excel defined name range and return a raw grid DataFrame.
    Keeps blank rows/columns so sheet structure is preserved for custom renderers.
    """
    try:
        dn = wb.defined_names[defined_name]
    except Exception as e:
        raise ValueError(f"Defined name '{defined_name}' not found.") from e

    destinations = list(dn.destinations)
    if not destinations:
        raise ValueError(f"Defined name '{defined_name}' has no destinations.")

    sheet_name, coord = destinations[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Defined name '{defined_name}' points to missing sheet '{sheet_name}'."
        )

    ws = wb[sheet_name]
    cell_range = ws[coord]

    if not isinstance(cell_range, tuple):
        return pd.DataFrame([[cell_range.value]])

    rows: List[List[object]] = []
    for row in cell_range:
        if isinstance(row, tuple):
            rows.append([cell.value for cell in row])
        else:
            rows.append([row.value])

    return pd.DataFrame(rows)


def load_named_table_from_bytes(
    xlsm_bytes: bytes,
    table_name: str,
    *,
    drop_empty_columns: bool = True,
) -> pd.DataFrame:
    """
    Load a named Excel table from workbook bytes by searching all sheets.
    """
    bio = BytesIO(xlsm_bytes)
    wb = load_workbook(bio, data_only=True)
    return _read_named_table_any_sheet(
        wb,
        table_name=table_name,
        drop_empty_columns=drop_empty_columns,
    )


def load_league_workbook_from_bytes(xlsm_bytes: bytes) -> ExcelLoadResult:
    """
    Load the league workbook from bytes and return key tables.
    data_only=True reads cached formula results saved by Excel.
    """
    bio = BytesIO(xlsm_bytes)
    wb = load_workbook(bio, data_only=True)

    # REQUIRED: fixtures table
    # Keep empty columns so schema is stable even before any results are entered.
    fixture_results = _read_named_table(
        wb,
        sheet_name="Fixture_Results",
        table_name="Fixture_Results_Table",
        drop_empty_columns=False,
    )

    # Optional tables
    league_table = None
    players = None
    teams = None
    league_data = None
    history_A_25_26 = None
    history_B_24_25 = None
    combined_stats = None
    top_performers = None

    # League table (pre-calculated in Excel)
    try:
        league_table = _read_named_table(
            wb,
            sheet_name="Fixture_Results",
            table_name="League_Table",
            drop_empty_columns=True,
        )
    except Exception:
        league_table = None

    try:
        players = _read_named_table(wb, sheet_name="Players", table_name="Player_Data", drop_empty_columns=True)
    except Exception:
        players = None

    try:
        teams = _read_named_table(wb, sheet_name="Teams", table_name="Teams_Table", drop_empty_columns=True)
    except Exception:
        teams = None

    try:
        league_data = _read_named_table(
            wb, sheet_name="League_Data", table_name="League_Data_Stats", drop_empty_columns=True
        )
    except Exception:
        league_data = None

    try:
        history_A_25_26 = _read_named_table_any_sheet(
            wb, table_name="A_25_26", drop_empty_columns=True
        )
    except Exception:
        history_A_25_26 = None

    try:
        history_B_24_25 = _read_named_table_any_sheet(
            wb, table_name="B_24_25", drop_empty_columns=True
        )
    except Exception:
        history_B_24_25 = None

    try:
        combined_stats = _read_named_table_any_sheet(
            wb, table_name="Combined_Stats", drop_empty_columns=True
        )
    except Exception:
        combined_stats = None

    try:
        top_performers = _read_defined_name_range(wb, "Top_Performers")
    except Exception:
        top_performers = None

    return ExcelLoadResult(
        fixture_results=fixture_results,
        league_table=league_table,
        players=players,
        teams=teams,
        league_data=league_data,
        history_A_25_26=history_A_25_26,
        history_B_24_25=history_B_24_25,
        combined_stats=combined_stats,
        top_performers=top_performers,
    )


def _canonical_col_name(name: str) -> str:
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _find_col_by_alias(df: pd.DataFrame, aliases: list[str]) -> str | None:
    by_canonical = {_canonical_col_name(c): c for c in df.columns}
    for alias in aliases:
        match = by_canonical.get(_canonical_col_name(alias))
        if match:
            return match
    return None


def load_week_stats_table_from_bytes(
    xlsm_bytes: bytes,
    week_number: int,
    *,
    drop_empty_columns: bool = True,
) -> pd.DataFrame | None:
    """
    Load a WeekNStats named table (e.g., Week1Stats) from workbook bytes.
    Returns None when the table is missing.
    """
    bio = BytesIO(xlsm_bytes)
    wb = load_workbook(bio, data_only=True)
    table_name = f"Week{int(week_number)}Stats"
    try:
        df = _read_named_table_any_sheet(
            wb,
            table_name=table_name,
            drop_empty_columns=drop_empty_columns,
        )
    except ValueError as err:
        if "not found in any worksheet" in str(err):
            return None
        raise
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load_week_stats_tables_from_bytes(
    xlsm_bytes: bytes,
    weeks: list[int] | None = None,
    *,
    drop_empty_columns: bool = True,
) -> dict[int, pd.DataFrame | None]:
    """
    Load WeekNStats tables for multiple weeks in one workbook pass.
    Missing tables are returned as None.
    """
    week_list = weeks if weeks is not None else list(range(1, 11))
    bio = BytesIO(xlsm_bytes)
    wb = load_workbook(bio, data_only=True)

    out: dict[int, pd.DataFrame | None] = {}
    for week in week_list:
        table_name = f"Week{int(week)}Stats"
        try:
            df = _read_named_table_any_sheet(
                wb,
                table_name=table_name,
                drop_empty_columns=drop_empty_columns,
            )
            df.columns = [str(c).strip() for c in df.columns]
            out[int(week)] = df
        except ValueError as err:
            if "not found in any worksheet" in str(err):
                out[int(week)] = None
            else:
                raise
    return out


def extract_week_fantasy_points_rows(
    week_df: pd.DataFrame,
    *,
    week_table_name: str,
) -> pd.DataFrame:
    """
    Normalize a WeekNStats DataFrame to player + fantasy points rows.
    Returns columns: player_key, player_id, player_name, fantasy_points.
    """
    if week_df is None or week_df.empty:
        return pd.DataFrame(columns=["player_key", "player_id", "player_name", "fantasy_points"])

    df = week_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    player_id_col = _find_col_by_alias(df, ["PlayerID", "Player Id", "Player ID"])
    player_name_col = _find_col_by_alias(df, ["Name", "Player", "Player Name"])
    fantasy_col = _find_col_by_alias(
        df,
        [
            "Fantasy Points",
            "FantasyPoints",
            "Total Fantasy Points",
            "Fantasy Points Total",
            "Fantasy Total Points",
            "Points",
            "Total Points",
            "Pts",
        ],
    )

    available_cols = ", ".join([str(c) for c in df.columns]) if len(df.columns) else "(none)"

    if fantasy_col is None:
        raise ValueError(
            f"{week_table_name} is missing a fantasy points column. Available columns: {available_cols}"
        )
    if player_id_col is None and player_name_col is None:
        raise ValueError(
            f"{week_table_name} is missing player identity columns. "
            f"Expected PlayerID or player name. Available columns: {available_cols}"
        )

    if player_id_col is not None:
        player_id = df[player_id_col].astype(str).str.strip()
        player_id = player_id.mask(player_id == "", "")
    else:
        player_id = pd.Series([""] * len(df), index=df.index, dtype="object")

    if player_name_col is not None:
        player_name = df[player_name_col].astype(str).str.strip()
        player_name = player_name.mask(player_name == "", "")
    else:
        player_name = pd.Series([""] * len(df), index=df.index, dtype="object")

    player_key = player_id.where(player_id != "", player_name)
    out = pd.DataFrame(
        {
            "player_key": player_key.astype(str).str.strip(),
            "player_id": player_id.astype(str).str.strip(),
            "player_name": player_name.astype(str).str.strip(),
            "fantasy_points": pd.to_numeric(df[fantasy_col], errors="coerce"),
        }
    )
    out = out[out["player_key"] != ""].copy()
    return out
